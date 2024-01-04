from django.shortcuts import render,redirect
from django.contrib import messages
from django.contrib.auth import authenticate,logout, login as dj_login
from django.shortcuts import get_object_or_404
from django.views import View
from .models import *
from accounts.models import *
from django.conf import settings
from django.http import JsonResponse
from django.utils import timezone
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from django.contrib.auth.hashers import make_password
from django.template.loader import render_to_string
from datetime import datetime, timedelta
from django.core.mail import send_mail
from django.urls import reverse
import tablib
from django.contrib.auth import get_user_model
import csv
from tablib import Dataset
from .resources import UserResource
from openpyxl import load_workbook
from io import BytesIO
from django.core.mail import EmailMessage
from dateutil.relativedelta import relativedelta
from django.db.models import Sum, F, ExpressionWrapper, Avg, Count
from collections import defaultdict
from django.db.models.functions import Coalesce
from django.db.models import Q
from django.db.models import Avg, F, Value
from django.http import HttpResponseServerError
from django.utils import timezone
from datetime import timedelta
from django.db.models.functions import ExtractMonth,ExtractYear, ExtractWeekDay
from django.db.models.functions import Coalesce
from django.db.models import Count, Subquery, OuterRef,Value
from django.core.exceptions import ObjectDoesNotExist

from django.db.models import Count, Case, When, Value, IntegerField, Sum
from django.db.models.functions import ExtractMonth, ExtractYear
from django.db.models import Q, F


def content_metrics():
    user_data = User.objects.filter(is_organization=True).annotate(book_count=Count('books'))
    content_metrics = [] 

    for user_instance in user_data:
        activate_user_count = User.objects.filter(
            institute=user_instance.institute,
            is_active=True,
            is_organization=False
        ).exclude(id=user_instance.id).count()

        dis = {
            'id': user_instance.id,
            'institute': user_instance.institute,
            'total_book': user_instance.book_count,
            'activate_user': activate_user_count
        }
        content_metrics.append(dis)

    # Sort content_metrics based on 'activate_user' in descending order
    content_metrics = sorted(content_metrics, key=lambda x: x['activate_user'], reverse=True)

    return content_metrics
def get_visit_count_and_day_name(date):
    day_name = date.strftime("%A")
    start_of_day = datetime.combine(date, datetime.min.time())
    end_of_day = start_of_day + timedelta(days=1)
    visit_count = BookVisit.objects.filter(start_time__range=[start_of_day, end_of_day]).count()
    
    return visit_count, day_name

def adoption_metrics():
    today = datetime.now().date()
    weekly_metrics = {}
    for day_of_week in range(1, 8):
        current_day = today - timedelta(days=(today.weekday() - day_of_week) % 7)
        visit_count, day_name = get_visit_count_and_day_name(current_day)
        weekly_metrics[day_name] = {
            'visit_count': visit_count,
            'day_name': day_name,
        }
    day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    sorted_weekly_metrics = {day: weekly_metrics[day] for day in day_order}

    return sorted_weekly_metrics
def corporate_users_non_corporate_users():
    users = User.objects.all().exclude(is_admin=True)
    total_users = users.count()

    if total_users == 0:
        # Handle the case where there are no users to avoid division by zero
        return 0, 0

    corporate_users = User.objects.filter(is_organization=False, is_admin=False, is_student=True).annotate(book_count=Count('books')).count()
    non_corporate_users = User.objects.filter(is_organization=False, is_admin=False, is_student=False).annotate(book_count=Count('books')).count()

    corporate_percentage = (corporate_users / total_users) * 100
    non_corporate_percentage = (non_corporate_users / total_users) * 100

    corporate_percentage = int(corporate_percentage)
    non_corporate_percentage = int(non_corporate_percentage)

    return corporate_percentage, non_corporate_percentage
def total_time_reading():
        # Get all book visits for the user
        book_visits = BookVisit.objects.all()

        total_time_per_book = defaultdict(int)
        for visit in book_visits:
            # Calculate time difference if start_time and end_time are present
            if visit.start_time and visit.end_time:
                time_difference = visit.end_time - visit.start_time
                total_time_per_book[visit.books_id] += time_difference.total_seconds()
def total_time_reading():
    # Get all book visits
    book_visits = BookVisit.objects.all()
    total_time_per_book = defaultdict(int)
    for visit in book_visits:
        # Calculate time difference if start_time and end_time are present
        if visit.start_time and visit.end_time:
            time_difference = visit.end_time - visit.start_time
            total_time_per_book[visit.books_id] += time_difference.total_seconds()
    return total_time_per_book
def per_week_average_time_spent():
    # Get all user sessions
    all_login_times = ManageLoginTime.objects.filter(
        login_start_time__isnull=False,
        login_end_time__isnull=False
    )

    # Calculate the total time spent by adding up the durations of all sessions
    total_duration = timedelta()

    for time in all_login_times:
        # Convert time objects to datetime objects
        start_datetime = datetime.combine(datetime.today(), time.login_start_time)
        end_datetime = datetime.combine(datetime.today(), time.login_end_time)

        session_duration = end_datetime - start_datetime
        total_duration += session_duration

    # Check if there are sessions before calculating averages
    if all_login_times.exists():
        # Calculate the total number of weeks (assuming 7 days per week)
        total_weeks = total_duration.days // 7 if total_duration.days else 1

        # Check if total_weeks is zero before attempting division
        if total_weeks > 0:
            # Calculate the overall average time spent per week by all users
            overall_average_time_per_week = total_duration / total_weeks

            # Convert the overall average time to a human-readable format (optional)
            overall_average_hours, remainder = divmod(overall_average_time_per_week.seconds, 3600)
            overall_average_minutes = remainder // 60

            return overall_average_hours, overall_average_minutes

    # Return a default value if there are no sessions
    return 0, 0
def reading_time():
    total_duration = BookVisit.objects.filter().aggregate(
    total_duration=Sum(F('end_time') - F('start_time'))
            )['total_duration']
    if total_duration is not None:
                # Now, total_duration is already a timedelta object
        hours, remainder = divmod(total_duration.seconds, 3600)
        minutes, _ = divmod(remainder, 60)
        total_time_admin_dashboard = f"Hours: {hours}, Minutes: {minutes}"
        return total_time_admin_dashboard
    else:
                # If total_duration is None, set default values
        hours, minutes = 0, 0
        total_time_admin_dashboard = f"Hours: {hours}, Minutes: {minutes}"
        return total_time_admin_dashboard
from django.db.models import DurationField


def dashboard(request):
    totol_servey = SurveyResponse.objects.count()
    total_time_per_book = total_time_reading()
    content_metric = content_metrics()
    sorted_weekly_metrics = adoption_metrics()
    total_time_admin_dashboard = reading_time()
    corporate_percentage,non_corporate_percentage  = corporate_users_non_corporate_users()
    last_month = datetime.now() - timedelta(days=30)
    users = User.objects.all().exclude(is_admin=True)
    active_users = User.objects.filter(is_active=True).exclude(is_admin=True).count()
    new_users_last_month = User.objects.filter(created_at__gte=last_month, is_admin=False, is_active=True).count()
    total_users = User.objects.all().exclude(is_admin=True)
    # corporate_users = User.objects.filter(institute__isnull=False).exclude(is_admin=True).exclude(is_organization=True).count()
    corporate_users = User.objects.filter(institute__isnull=False,is_admin=False,is_organization=False,is_staff=False).count()
    non_corporate_users = User.objects.filter(is_student=False).exclude(is_organization=True).exclude(is_admin=True).count()
    year = current_year = datetime.now().year
    all_months_years = User.objects.annotate(
    month=ExtractMonth('created_at'),
    year=ExtractYear('created_at')
    ).values('year', 'month').distinct()
    users_by_month_year = all_months_years.annotate(
        total_users=Coalesce(Count('id', distinct=True), Value(0)),
        activated_users=Coalesce(Count('id', filter=Q(is_active=True), distinct=True), Value(0))
    )
    user_growth_rate = 0
    if active_users > 0:
        user_growth_rate = round(((new_users_last_month - 0) / active_users) * 100)
    total_user = users.count()
    total_contactus = ContactUs.objects.all().count()
    total_book = Books.objects.filter(is_active=True).count()
    total_collection = Collections.objects.all().count()
    total_category = BookCategory.objects.all().count()
    total_new_user = User.objects.filter(is_superuser=False, is_staff=False).order_by('id')[:10].count()
    total_coprorate = User.objects.filter(is_organization=True,).order_by('id')[:10].count()
    user_visits = UserVisit.objects.all()
    total_visits = user_visits.count()
    total_staff = User.objects.filter(is_staff=True).exclude(is_admin=True).count()
    total_visit = BookVisit.objects.all().count()
    total_confirmed = User.objects.filter(is_active=True).count()
    total_unconfirmed = User.objects.filter(is_active=False, is_admin=False).count()
    if total_confirmed > 0 and total_unconfirmed > 0:
        conversion_rate = (total_confirmed / total_unconfirmed) * 100
        conversion_rate = min(100, conversion_rate)
        # Format the conversion_rate with two decimal places
        conversion_rate = f"{conversion_rate:.2f}"
    else:
        conversion_rate = 0  # Set conversion_rate to 0 when either total_confirmed or total_unconfirmed is not greater than 0


    h , m = average_time_spent_admin_dashboard()
    overall_average_hours, o = per_week_average_time_spent()
    users_avg_time = f"{h} H, {m} Min."
    total_avg_h = f"{h}"
    top_institutes = (
        BookVisit.objects
        .exclude(institute_name__isnull=True)
        .values('institute_name')
        .annotate(
            total_duration=Coalesce(Sum(F('end_time') - F('start_time')), timedelta(seconds=0), output_field=DurationField())
        )
        .order_by('-total_duration')[:10]
    )

    # Fetching additional details for each institute
    for entry in top_institutes:
        institute_name = entry['institute_name']
        
        # Fetching the first user for the institute
        user = BookVisit.objects.filter(institute_name=institute_name).first()

        if user:
            user_id = User.objects.get(id=user.user.added_user)
            try:
                user_info = User.objects.get(id=user_id.id, is_organization=True)
                books_count = user_info.books.count()
                entry['books_count'] = books_count
            except User.DoesNotExist:
                books_count = 0
            # Converting total_duration to hours and minutes or setting it to 0 if it's None
            total_duration = entry['total_duration'] or timedelta(seconds=0)
            total_duration_seconds = total_duration.total_seconds()
            hours = int(total_duration_seconds // 3600)
            minutes = int((total_duration_seconds % 3600) // 60)
            entry['users_avg_time'] = f"{hours} H, {minutes} Min."
        else:
            # If no user found, set default values
            entry['books_count'] = 0
            entry['users_avg_time'] = "0 H, 0 Min."
    context = {
        'active_users':active_users,
        'total_visits': total_visits,
        'total_user':total_user,
        'total_book':total_book,
        'total_contactus':total_contactus,
        'total_collection':total_collection,
        'total_category':total_category,
        'total_new_user':total_new_user,
        'total_coprorate':total_coprorate,
        'conversion_rate':conversion_rate,
        'total_staff':total_staff,
        'active21': 'active12',
        'user_growth_rate':user_growth_rate,
        'users_avg_time': users_avg_time,
        'corporate_users':corporate_users,
        'non_corporate_users':non_corporate_users,
        'users_by_month_year':users_by_month_year,
        'year':year,
        'top_institutions': top_institutes,
        'content_metrics': content_metric,
        'sorted_weekly_metrics': sorted_weekly_metrics,
        'total_visit':total_visit,
        'corporate_percentage': corporate_percentage,
        'non_corporate_percentage' : non_corporate_percentage,
        'total_avg_h':total_avg_h,
        'overall_average_hours':overall_average_hours,
        'totol_servey':totol_servey,
        'total_time_admin_dashboard':total_time_admin_dashboard
        

        
    }
    if request.user.is_authenticated:
        if request.user.is_superuser:
            return render(request, 'dashboard.html', {**context})
        elif hasattr(request.user, 'is_organization') and request.user.is_organization:
            return redirect('Dashboardorgani')


######### login views start here ########
def login(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            return redirect('dashboard')
        elif hasattr(request.user, 'is_organization') and request.user.is_organization:
            return redirect('Dashboardorgani')

    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        user = authenticate(request=request, email=email, password=password)
        if user is not None:
            if hasattr(user, 'is_superuser') and user.is_superuser:
                dj_login(request, user)
                return redirect('dashboard')
            elif hasattr(user, 'is_organization') and user.is_organization:
                login_time = datetime.now().strftime('%H:%M:%S')
                ManageLoginTime.objects.create(users=user, login_start_time=login_time)
                dj_login(request, user)
                return redirect('Dashboardorgani')
            else:
                messages.error(request, 'You Are Not Admin User')
        else:
            messages.error(request, 'Invalid Email or Password')

    return render(request, 'admin_login.html')




class ProfileView(View):
    def get(self,request,id):
        data = User.objects.get(id=id)
        book_count = Books.objects.filter(is_active=True).count()
        book_visit = BookVisit.objects.all().count()
        read_book = BookReadProgress.objects.filter(user=data,book__is_active=True).count()
        return render(request,'view_profile.html',{'data':data,'book_visit':book_visit,'read_book':read_book})



class SendNotificationSingle(View):
    def get(self,request,id):
        data = Notifications.objects.filter(id=id)
        user = User.objects.get(id=id)
        return render(request,'single_sendnotification.html',{'data':data,'user':user})
    def post(self,request,id):
        receiver = request.POST.get('receiver')
        receiver_email = User.objects.get(email=receiver)
        subject = request.POST.get('subject')
        message = request.POST.get('message')
        Notifications.objects.create(receiver=receiver_email,subject=subject,message=message)
        return redirect('detailsnotification')









class BookInvitationSend(View):
    def get(self, request):
        user = User.objects.filter(is_superuser=False,is_staff=False)

        book = Books.objects.all()  # Make sure you import your Book model
        return render(request, 'add_invitation.html', {'user': user, 'book': book})
    def post(self, request):
        invited_users = request.POST.getlist('invited_user')
        title = request.POST.get('title')
        book_title = request.POST.get('book')
        for invited_user in invited_users:
            invited_user = invited_user.strip()
            if invited_user:
                user = User.objects.filter(email=invited_user).first()
                if user:
                    book_instance = Books.objects.get(title=book_title)
                    invitation = BookInvitation.objects.create(book=book_instance, invited_user=user, title=title)
                    subject = 'You have received a book invitation'
                    sub = 'click on the link to read new book' 
                    link = 'http://64.227.157.173:8002/'
                    message = f'You have received an invitation to join the book "{book_instance.title}" with the title "{title}".\n\n{sub} \n\n{link}'
                    from_email = 'your@gmail.com'
                    recipient_list = [invited_user]
                    send_mail(subject, message, from_email, recipient_list)
        return redirect('book_invitation')

class BookInvitationDetails(View):
    def get(self,request):
        data = BookInvitation.objects.all()
        return render(request,'details_invitation.html',{'data':data})

class BookInvitationDelete(View):
    def get(self,request,id):
        data = BookInvitation.objects.get(id=id)
        data.delete()
        return redirect('book_invitation')


######## admin update profile ########
class AdminUpdateprofile(View):
    def get(self, request):
        if request.user.is_authenticated:
            data=User.objects.filter(id=request.user.id,is_superuser=True)
            return render(request, 'profile_setting.html',{'data':data})
        return redirect('login')
    def post(self, request):
        id = request.user.id
        usern = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        phone_number = request.POST.get('phone_number')
        profile_pic = request.FILES.get('profile_pic')
        if profile_pic is not None:
            id = request.user.id
            user = User.objects.get(id=id)
            user.first_name = usern
            user.email = email
            user.last_name = last_name
            user.phone_number = phone_number
            user.profile_pic = profile_pic
            user.save()
            messages.success(request,'User Profile Updated Successfully!')
            return redirect('updateprofile')
        else:
            user = User.objects.get(id=id)
            user.first_name = usern
            user.email = email
            user.last_name = last_name
            user.phone_number = phone_number
            user.profile_pic = user.profile_pic
            user.save()
            messages.success(request,'User Profile Updated Successfully!!')
            return redirect('updateprofile')





# class ChangePassword(View):
#     def post(self,request):
#         old_password = request.POST.get('old_pass')

#         new_password = request.POST.get('new_pass')
#         confirm_password = request.POST.get('confirm_new_pass')
#         user = User.objects.get(id=request.user.id)
#         if old_password and new_password and confirm_password:
#             if user.check_password(old_password):
#                 if new_password == confirm_password:
#                     user.set_password(new_password)
#                     user.save()
#                     data = {'msg': 'Password changed successfully.'}
#                     return JsonResponse(data)
#                 else:
#                     data = {'msg': 'New password and confirm password do not match.'}
#             else:
#                 data = {'msg': 'Old password is incorrect.'}
#         else:
#             data = {'msg': 'All fields are required.'}
#         return JsonResponse(data)


from django.contrib.auth import update_session_auth_hash

class ChangePassword(View):
    def post(self, request):
        old_password = request.POST.get('old_password')
        password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        if old_password and password and confirm_password:
            if request.user.check_password(old_password):
                if password == confirm_password:
                    user = User.objects.get(id=request.user.id)
                    user.set_password(password)
                    user.save()
                    update_session_auth_hash(request, user)
                    messages.success(request, 'Password changed successfully.')
                    return redirect('dashboard')  
                else:
                    messages.error(request, 'New password and confirm password do not match.')
            else:
                messages.error(request, 'Old password is incorrect.')
        else:
            messages.error(request, 'All fields are required.')

        return redirect('updateprofile')




################# admin logout views ##########
def logout_view(request):
    if request.user.is_superuser:
        logout(request)
        return redirect('login')
    elif request.user.is_staff:
        logout(request)
        return redirect('stafflogin')
    elif request.user.is_organization:
        user_id = request.user.id
        user = User.objects.get(id=user_id)
        user_login_time = ManageLoginTime.objects.filter(users=user).last()
        logout_time = datetime.now().strftime('%H:%M:%S')
        user_login_time.login_end_time = logout_time
        user_login_time.save()
        logout(request)
        return redirect('login')
    else:
        logout(request)
        return redirect('login')



################## add organizations views #######
class Add_Organizations(View):
    def get(self, request):
        data = User.objects.all()
        return render(request,'add_organization.html',{'data':data})
    def post(self, request):
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        phone_number = request.POST.get('phone_number')
        password = request.POST.get('password')
        if User.objects.filter(email=email).exists():
            messages.error(request, 'Email id  Already Exist...? Use Diffrent Email id to Create Account.')
        else:
            user =User.objects.create_user(first_name=first_name,last_name=last_name, email=email, phone_number=phone_number,password=password)
            user.is_organization=True
            user.save()
            messages.success(request,'Organizations Registration Successfully..!!')
        return redirect('user_details')



######## Details organization views ########
class Details_User(View):
    def get(self, request):
        data = User.objects.all().exclude(is_superuser=True).exclude(is_staff=True)
        filter_corporate = User.objects.filter(is_organization=True)
        filter_non_corporate = User.objects.filter(is_organization=False)
        active = User.objects.filter(is_active=True)
        inactive = User.objects.filter(is_active=False)
        institute = User.objects.values_list('institute', flat=True).distinct().exclude(is_superuser=True).exclude(is_staff=True).exclude(is_enduser=True)
        filtering_options = {
            'all': 'All',
            'corporate': 'Corporate',
            'non_corporate': 'Non-Corporate',
            
        }

        return render(request, 'manage_user.html', {
            'data': data,
            'filter_corporate': filter_corporate,
            'filter_non_corporate': filter_non_corporate,
            'filtering_options': filtering_options,
            'active_users': active, 
            'inactive_users': inactive, 
            'institutes': institute,
        })

class Delete_User(View):
    def get(self,request,id):
        data = User.objects.get(id=id)
        data.delete()
        messages.success(request,'User Delete Successfully')
        return redirect('user_details')






class Details_Enduser(View):
    def get(self,request):
        data = User.objects.all().exclude(is_organization=True)
        return render(request,'manage_enduser.html',{'data':data})


class Details_DemoRequest(View):
    def get(self,request):
        data = RequestDemo.objects.all()
        return render(request,'manage_requestdemo.html',{'data':data})


class Deletedemorequest(View):
    def get(self,request,id):
        data = RequestDemo.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully')
        return redirect('Demo_request')



class Active_inActiveUser(View):
    def post(self, request):
        user_id = request.POST['id']
        user = User.objects.get(id=user_id)
        if user.is_active is False:
            user.is_active = True
            user.save()
            return redirect('organization_details')
        elif user.is_active is True:
            user.is_active = False
            user.save()
            return redirect('organization_details')
        else:
            return HttpResponse("User Not Valid")



########  ######## 
class AddBookCategory(View):
    def get(Self,request):
        data = BookCategory.objects.all()
        return render(request,'add_category.html',{'data':data})
    def post(Self,request):
        book_category = request.POST.get('book_category')
        BookCategory.objects.create(book_category=book_category)
        messages.success(request, "Category Added Successfully..!!")
        return redirect('category_details')
    

class Category_details(View):
    def get(self,request):
        data = BookCategory.objects.all()
        return render(request,'category_details.html',{'data':data,'active24':'active12'})


class Category_Edit(View):
    def get(self, request, id):
        data = BookCategory.objects.get(id=id)
        return render(request, 'edit_category.html', {'data': data})
    def post(self, request, id):
        data = BookCategory.objects.get(id=id)
        book_category = request.POST.get('book_category')
        if data.created_at is None:
            data.created_at = timezone.now().date()
        data.book_category = book_category
        data.save()
        messages.success(request, "Category Updated Successfully..!!")
        return redirect('category_details')


class Category_Delete(View):
    def get(self,request,id):
        data = BookCategory.objects.get(id=id)
        data.delete()
        messages.success(request,'Category Delete Successfully!..')
        return redirect('category_details')


########### add book views ###########
class Add_book(View):
    def get(self, request):
        category = BookCategory.objects.all()
        return render(request, 'add_books.html', {'category': category})
    def post(self, request):
        # heading = request.POST.get('heading')
        description = request.POST.get('description')
        title = request.POST.get('title')
        author = request.POST.get('author')
        image = request.FILES.get('image')
        book_pdf = request.FILES.get('book_pdf')
        price = request.POST.get('price') or 0
        publication_date = request.POST.get('publication_date')
        category = request.POST.get('book_category')
        # pages = request.POST.get('pages')
        collection = request.POST.get('collection')
        category = get_object_or_404(BookCategory, book_category=category)
        Books.objects.create(
            category=category,
            description=description,
            title=title,
            author=author,
            image=image,
            book_pdf=book_pdf,
            price=price,
            publication_date=publication_date,
            # pages=pages,
            collection=collection
        )
        messages.success(request, "Book Added Successfully..!!")
        return redirect('details_book')




######### book details views ########
class Details_Book(View):
    def get(self,request):
        data = Books.objects.all()
        select_categories = BookCategory.objects.all()
        return render(request,'details_books.html',{'data':data,'select_categories':select_categories,'active26':'active12'})

class ActivateAccountView(View):
    def get(self, request, id):
        print('hello ajay')
        user = User.objects.get(id=id)
        user.is_active = True
        user.save()
        messages.success(request, 'Your account has been activated. You can now log in.')
        return redirect('enduser-login') 


########### book edit views #######
class Edit_book(View):
    def get(self,request,id):
        data = Books.objects.get(id=id)
        category = BookCategory.objects.all()
        return render(request,'edit_books.html',{'data':data,'category':category})
    def post(self, request, id):
        data = Books.objects.get(id=id)
        category = request.POST.get('book_category')
        description = request.POST.get('description')
        title = request.POST.get('title')
        author = request.POST.get('author')
        image = request.FILES.get('image')
        book_pdf = request.FILES.get('book_pdf')
        price = request.POST.get('price')
        publication_date = request.POST.get('publication_date')
        category = request.POST.get('book_category')
        # pages = request.POST.get('pages')
        collection = request.POST.get('collection')
        category = get_object_or_404(BookCategory, book_category=category)
        if image and book_pdf is not None:
            data = Books(id=id,collection=collection,description=description,author=author,publication_date=publication_date,book_pdf=book_pdf,title=title,cimage=image,category=category,price=price).save()
        elif image is not None:
            data = Books(id=id,collection=collection,description=description,author=author,publication_date=publication_date,book_pdf=data.book_pdf,title=title,image=image,category=category,price=price).save()
        elif book_pdf is not None:
            data = Books(id=id,collection=collection,description=description,author=author,publication_date=publication_date,book_pdf=book_pdf,title=title,image=data.image,category=category,price=price).save()
        else:
            data = Books(id=id,collection=collection,description=description,author=author,publication_date=publication_date,title=title, book_pdf=data.book_pdf,image=data.image,category=category,price=price).save()
        return redirect('details_book')
       



############# delete book views #####
class Delete_Book(View):
    def get(self,request,id):
        data = Books.objects.get(id=id)
        data.delete()
        messages.success(request, "Data Delete Successfully..!!")
        return redirect('details_book')
    



### active and inactive 
class Active_inActive(View):
    def post(self, request):
        user_id = request.POST['id']
        user = Books.objects.get(id=user_id)
        if user.is_active is False:
            user.is_active = True
            user.save()
            return redirect('details_book')
        elif user.is_active is True:
            user.is_active = False
            user.save()
            return redirect('details_book')
        else:
            return HttpResponse("User Not Valid")

######### contact us views start here ######
class ContactUsView(View):
    def get(self,request):
        data = ContactUs.objects.all()
        return render(request,'contact_details.html',{'data':data})

class ContactUs_Delete(View):
    def get(self,request,id):
        data = ContactUs.objects.get(id=id)
        data.delete()
        messages.success(request,'Contact Us Deleted Successfully..!!')
        return redirect('messageus')
    

class AddContactBanner(View):
    def get(self,request):
        data = ContactBanner.objects.all()
        return render(request,'add_contactbanner.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        description = request.POST.get('description')
        image = request.FILES.get('image')
        ContactBanner.objects.create(heading=heading,image=image,description=description)
        return redirect('ContactusBanner_details')

class DetailsContactBanner(View):
    def get(self,request):
        data = ContactBanner.objects.all()
        return render(request,'details_contactbanner.html',{'data':data})
    
class EditContactBanner(View):
    def get(self,request,id):
        data = ContactBanner.objects.get(id=id)
        return render(request,'edit_contactbanner.html',{'data':data})
    def post(self,request,id):
        data = ContactBanner.objects.get(id=id)
        heading = request.POST.get('heading')
        description = request.POST.get('description')
        image = request.FILES.get('image')
        if image is not None:
            ContactBanner(id=id,heading=heading,image=image,description=description).save()
        else:
            ContactBanner(id=id,heading=heading,image=data.image,description=description).save()
        return redirect('ContactusBanner_details')


class DeleteContactBanner(View):
    def get(self,request,id):
        data = ContactBanner.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('ContactusBanner_details')



class AddGetInTouch(View):
    def get(self,request):
        data = GetInTouch.objects.all()
        return render(request,'add_getintouch.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        description = request.POST.get('description')
        address = request.POST.get('address')
        email = request.POST.get('email')
        call = request.POST.get('call')
        GetInTouch.objects.create(description=description,heading=heading,address=address,email=email,call=call)
        return redirect('Getintouch_details')
    
class DetailGetInTouch(View):
    def get(self,request):
        data = GetInTouch.objects.all()
        return render(request,'detail_getintouch.html',{"data":data})
        
class DeleteGetInTouch(View):
    def get(self,request,id):
        data = GetInTouch.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('Getintouch_details')



class EditGetInTouch(View):
    def get(self,request,id):
        data = GetInTouch.objects.get(id=id)
        return render(request,'edit_getintouch.html',{'data':data})
    def post(self,request,id):
        data = GetInTouch.objects.get(id=id)
        heading = request.POST.get('heading')
        description = request.POST.get('description')
        address = request.POST.get('address')
        email = request.POST.get('email')
        call = request.POST.get('call')
        GetInTouch(id=id,description=description,heading=heading,address=address,email=email,call=call).save()
        return redirect('Getintouch_details')


################ home page views start here.. #######
class AddHome(View):
    def get(self,request):
        data = Home.objects.all()
        return render(request,'add_home.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        description = request.POST.get('description')
        image = request.FILES.get('image')
        youtube_link = request.POST.get('youtube_link')
        Home.objects.create(heading=heading,description=description,image=image,youtube_link=youtube_link)
        return redirect('home_details')
    


class HomeDetails(View):
    def get(self,request):
        data = Home.objects.all()
        return render(request,'details_home.html',{'data':data})
    

class EditHome(View):
    def get(self,request,id):
        data = Home.objects.get(id=id)
        return render(request,'edit_home.html',{'data':data})
    def post(self,request,id):
        data = Home.objects.get(id=id)
        heading = request.POST.get('heading')
        description = request.POST.get('description')
        image = request.FILES.get('image')
        youtube_link = request.POST.get('youtube_link')
        if image is not None:
            Home(id=id,heading=heading,description=description,image=image,youtube_link=youtube_link).save()
        else:
            Home(id=id,heading=heading,description=description,image=data.image,youtube_link=youtube_link).save()
        return redirect('home_details')


class HomeDelete(View):
    def get(self,request,id):
        data = Home.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Delete Successfully!..')
        return redirect('home_details')
    
################# learn everthing views start here########
class AddLearn(View):
    def get(self,request):
        data = LearnEverything.objects.all()
        return render(request,'add_learneverthing.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        description = request.POST.get('description')
        image = request.FILES.get('image')
        LearnEverything.objects.create(heading=heading,description=description,image=image)
        return redirect('learn_details')
    


class LearnDetails(View):
    def get(self,request):
        data = LearnEverything.objects.all()
        return render(request,'detail_learneverthing.html',{'data':data})
    

class EditLearn(View):
    def get(self,request,id):
        data = LearnEverything.objects.get(id=id)
        return render(request,'edit_learneverthing.html',{'data':data})
    def post(self,request,id):
        data = LearnEverything.objects.get(id=id)
        heading = request.POST.get('heading')
        description = request.POST.get('description')
        image = request.FILES.get('image')
        if image is not None:
            LearnEverything(id=id,heading=heading,description=description,image=image).save()
        else:
            LearnEverything(id=id,heading=heading,description=description,image=data.image).save()
        return redirect('learn_details')


class LearnDelete(View):
    def get(self,request,id):
        data = LearnEverything.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Delete Successfully!..')
        return redirect('learn_details')
    
################## how does views start  here ##########################
class AddHowDoes(View):
    def get(self,request):
        data = HowDoesebooks.objects.all()
        return render(request,'add_howdoese.html',{'data':data})
    def post(self,request):
        top_heading = request.POST.get('top_heading')
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        HowDoesebooks.objects.create(top_heading=top_heading,heading=heading,image=image,description=description)
        return redirect('howdoes_details')




class Details_HowDoes(View):
    def get(self,request):
        data = HowDoesebooks.objects.all()
        return render(request,'details_howdoes.html',{"data":data})


class EditHowDoes(View):
    def get(self,request,id):
        data = HowDoesebooks.objects.get(id=id)
        return render(request,'edit_howdoes.html',{'data':data})
    def post(self,request,id):
        data = HowDoesebooks.objects.get(id=id)
        top_heading = request.POST.get('top_heading')
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        if image is not None:
            HowDoesebooks(id=id,top_heading=top_heading,heading=heading,image=image,description=description).save()
        else:
            HowDoesebooks(id=id,top_heading=top_heading,heading=heading,image=data.image,description=description).save()

        return redirect('howdoes_details')


class DeleteHowDoes(View):
    def get(self,request,id):
        data = HowDoesebooks.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Delete successfully!..' )
        return redirect('howdoes_details')
    


class AddEbookExp(View):
    def get(self,request):
        data = EBookExperience.objects.all()
        return render(request,'add_bookExp.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        title = request.POST.get('title')
        image1 = request.FILES.get('image1')
        image2 = request.FILES.get('image2')
        description = request.POST.get('description')
        sub_title = request.POST.get('sub_title')
        icon = request.FILES.get('icon')
        EBookExperience.objects.create(heading=heading,title=title,image1=image1,image2=image2,description=description,
        sub_title=sub_title,icon=icon)
        return redirect('ebookexp_details')



class Details_EbookExp(View):
    def get(self,request):
        data = EBookExperience.objects.all()
        return render(request,'details_bookExp.html',{'data':data})

class Delete_EbookExp(View):
    def get(self,request,id):
        data = EBookExperience.objects.get(id=id)
        data.delete()
        return redirect('ebookexp_details')


class Edit_EbookExp(View):
    def get(self,request,id):
        data = EBookExperience.objects.get(id=id)
        return render(request,'edit_bookexp.html',{'data':data})
    def post(self,request,id):
        data = EBookExperience.objects.get(id=id)
        heading = request.POST.get('heading')
        title = request.POST.get('title')
        image1 = request.FILES.get('image1')
        image2 = request.FILES.get('image2')
        description = request.POST.get('description')
        sub_title = request.POST.get('sub_title')
        icon = request.FILES.get('icon')
        data.heading = heading
        data.title = title
        data.description = description
        data.sub_title = sub_title
        if image1:
            data.image1 = image1
        if image2:
            data.image2 = image2
        if icon:
            data.icon = icon
        data.save()
        return redirect('ebookexp_details')



class AddOurService(View):
    def get(self,request):
        data = OurServices.objects.all()
        return render(request,'add_ourservice.html',{'data':data})
    def post(self,request):
        top_heading = request.POST.get('top_heading')
        heading = request.POST.get('heading')
        icon = request.POST.get('icon')
        sub_heading = request.POST.get('sub_heading')
        description = request.POST.get('description')
        background_image = request.FILES.get('background_image')
        OurServices.objects.create(top_heading=top_heading,heading=heading,icon=icon,
        sub_heading=sub_heading,description=description,background_image=background_image)
        return redirect('ourservice_details')


class Details_Service(View):
    def get(self,request):
        data = OurServices.objects.all()
        return render(request,'details_ourservice.html',{'data':data})
    


class Edit_Ourservice(View):
    def get(self,request,id):
        data = OurServices.objects.get(id=id)
        return render(request,'edit_ourservice.html',{'data':data})
    def post(self,request,id):
        data = OurServices.objects.get(id=id)
        top_heading = request.POST.get('top_heading')
        heading = request.POST.get('heading')
        icon = request.POST.get('icon')
        sub_heading = request.POST.get('sub_heading')
        description = request.POST.get('description')
        background_image = request.FILES.get('background_image')
        if background_image is not None:
            OurServices(id=id,top_heading=top_heading,heading=heading,icon=icon,
            sub_heading=sub_heading,description=description,background_image=background_image).save()
        else:
            OurServices(id=id,top_heading=top_heading,heading=heading,icon=icon,
            sub_heading=sub_heading,description=description,background_image=data.background_image).save()
        return redirect('ourservice_details')



class Delete_Service(View):
    def get(self,request,id):
        data = OurServices.objects.get(id=id)
        data.delete()
        return redirect('ourservice_details')
    


class AddOur_Partner(View):
    def get(self,request):
        data = OurPartner.objects.all()
        return render(request,'add_ourpartner.html',{'data':data})
    def post(self,request):
        image = request.FILES.get('image')
        OurPartner.objects.create(image=image)
        return redirect('ourpartner_details')


class Details_OurPartner(View):
    def get(self,request):
        data = OurPartner.objects.all()
        return render(request,'details_ourpartner.html',{'data':data})
    

class Delete_OurPartner(View):
    def get(self,request,id):
        data = OurPartner.objects.get(id=id)
        data.delete()
        return redirect('ourpartner_details')
    

class EditOur_Partner(View):
    def get(self,request,id):
        data = OurPartner.objects.get(id=id)
        return render(request,'edit_ourpartner.html',{'data':data})
    def post(self,request,id):
        data = OurPartner.objects.get(id=id)
        image = request.FILES.get('image')
        if image is not None:
            OurPartner(id=id,image=image).save()
        else:
            OurPartner(id=id,image=data.image).save()
        return redirect('ourpartner_details')
    




class AddStudio(View):
    def get(self,request):
        data = StudioExperience.objects.all()
        return render(request,'add_studio.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        title = request.POST.get('title')
        image1 = request.FILES.get('image1')
        image2 = request.FILES.get('image2')
        description = request.POST.get('description')
        icon_title = request.POST.get('icon_title')
        icon = request.FILES.get('icon')
        StudioExperience.objects.create(heading=heading,title=title,image1=image1,image2=image2,description=description
        ,icon_title=icon_title,icon=icon)
        return redirect('studio_details')


class Details_studio(View):
    def get(self,request):
        data = StudioExperience.objects.all()
        return render(request,'details_studio.html',{'data':data})


class Delete_studio(View):
    def get(self,request,id):
        data = StudioExperience.objects.get(id=id)
        data.delete()
        return redirect('studio_details')

class Edit_Studio(View):
    def get(self,request,id):
        data = StudioExperience.objects.get(id=id)
        return render(request,'edit_studio.html',{'data':data})
    def post(self,request,id):
        data = StudioExperience.objects.get(id=id)
        heading = request.POST.get('heading')
        title = request.POST.get('title')
        image1 = request.FILES.get('image1')
        image2 = request.FILES.get('image2')
        description = request.POST.get('description')
        icon_title = request.POST.get('icon_title')
        icon = request.FILES.get('icon')
        if image1 is not None:
            data.image1=image1
            data.save()
        if image2 is not None:
            data.image2=image2
            data.save()
        if icon:
            data.icon=icon
            data.heading=heading
            data.description=description
            data.icon_title=icon_title
            data.save()
        return redirect('studio_details')
    


####### footer views start here ##
class Add_Footer(View):
    def get(self,request):
        data = Footer.objects.all()
        return render(request,'add_footer.html',{'data':data})
    def post(self,request):
        facebook = request.POST.get('facebook')
        instagram = request.POST.get('instagram')
        twitter = request.POST.get('twitter')
        description = request.POST.get('description')
        logo = request.FILES.get('logo')
        linkedin = request.POST.get('linkedin')
        Footer.objects.create(linkedin=linkedin,logo=logo,facebook=facebook,instagram=instagram,twitter=twitter,description=description)
        return redirect('footer_details')

class Details_footer(View):
    def get(self,request):
        data = Footer.objects.all()
        return render(request,'details_footer.html',{'data':data})

class EditFooter(View):
    def get(self,request,id):
        data = Footer.objects.get(id=id)
        return render(request,'edit_footer.html',{'data':data})
    def post(self,request,id):
        data = Footer.objects.get(id=id)
        logo = request.FILES.get('logo')
        facebook = request.POST.get('facebook')
        instagram = request.POST.get('instagram')
        twitter = request.POST.get('twitter')
        description = request.POST.get('description')
        linkedin = request.POST.get('linkedin')
        if logo is not None:
            Footer(id=id,linkedin=linkedin,logo=logo,facebook=facebook,instagram=instagram,twitter=twitter,description=description).save()
        else:
            Footer(id=id,linkedin=linkedin,logo=data.logo,facebook=facebook,instagram=instagram,twitter=twitter,description=description).save()
        return redirect('footer_details')

class DeleteFooter(View):
    def get(self,request,id):
        data=Footer.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('footer_details')


########## #############
# class AddCollection(View):
#     def get(self, request):
#         data = Books.objects.all()
#         return render(request, 'add_collection.html', {'data': data})
#     def post(self, request):
        # name = request.POST.get('name')
        # image = request.FILES.get('image')
        # books = request.POST.getlist('books')
        # data = Collections.objects.create(name=name,image=image)
        # for book_id in books:
        #     book = Books.objects.get(id=book_id)
        #     CollectionsList.objects.create(Collections_name=data,books_id=book)
        # return redirect('collection_details')
class AddCollection(View):
    def get(self, request):
        data = Books.objects.filter(is_active=True)
        selected_books = Collections.objects.all().values_list('id', flat=True)
        return render(request, 'add_collection.html', {'data': data,'selected_books':selected_books})
    def post(self, request):
        name = request.POST.get('name')
        image = request.FILES.get('image')
        books = request.POST.getlist('books')
        data = Collections.objects.create(name=name, image=image)

        for book_id in books:
            book = Books.objects.get(id=book_id)
            CollectionsList.objects.create(Collections_name=data, books_id=book)

        return redirect('collection_details')
class EditCollection(View):
    def get(self, request, id):
        book = Books.objects.filter(is_active=True)
        data = Collections.objects.get(id=id)
        return render(request, 'edit_collection.html', {'data': data,'book':book})
    def post(self, request, id):
        data = Collections.objects.get(id=id)
        name = request.POST.get('name')
        image = request.FILES.get('image')
        books = request.POST.getlist('books')
        existing_collection = CollectionsList.objects.filter(Collections_name=data, books_id__in=books).exists()
        if existing_collection:
            messages.error(request, 'These books are already added in the collection.')
            return redirect('collection_details')
        data.name=name
        if image:
           data.image = image
        data.save()
        for book_id in books:
            book = Books.objects.get(id=book_id)
            CollectionsList.objects.create(Collections_name=data, books_id=book)
        return redirect('collection_details')

class CollectionDetails(View):
    def get(self, request):
        data = Collections.objects.all()
        return render(request, 'collection_details.html', {'data': data,'active25':'active12'})


class CollectionDelete(View):
    def get(self,request,id):
        data = Collections.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('collection_details')
    



class CollectionView(View):
    def get(self,request,id):
        data = CollectionsList.objects.filter(Collections_name=Collections.objects.get(id=id))
        return render(request,'collection_view.html',{'data':data})
    


class CollectionViewDelete(View):
    def get(self, request, id):
        data = CollectionsList.objects.get(id=id)
        data.delete()
        return redirect('collection_details')
    


######## about us views start here ####
class AddAboutBanner(View):
    def get(self,request):
        data = AboutBanner.objects.all()
        return render(request,'add_aboutbanner.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        banner_image = request.FILES.get('banner_image')
        description = request.POST.get('description')
        AboutBanner.objects.create(heading=heading,banner_image=banner_image,description=description)
        return redirect('AboutBanner_details')
    

class DetailsAbout(View):
    def get(self,request):
        data = AboutBanner.objects.all()
        return render(request,'details_Aboutbanner.html',{'data':data})
    

class DeleteAbout(View):
    def get(self,request,id):
        data = AboutBanner.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('AboutBanner_details')
    

class EditAboutBanner(View):
    def get(self,request,id):
        data = AboutBanner.objects.get(id=id)
        return render(request,'edit_aboutbanner.html',{'data':data})
    def post(self,request,id):
        data = AboutBanner.objects.get(id=id)
        heading = request.POST.get('heading')
        banner_image = request.FILES.get('banner_image')
        description = request.POST.get('description')
        if banner_image is not None:
            AboutBanner(id=id,heading=heading,banner_image=banner_image,description=description).save()
        else:
            AboutBanner(id=id,heading=heading,banner_image=data.banner_image,description=description).save()
        return redirect('AboutBanner_details')




############## the foundation ##########
class AddTheFoundation(View):
    def get(self,request):
        data = TheFoundation.objects.all()
        return render(request,'add_thefoundation.html',{'data':data})
    def post(self,request):
        top_heading = request.POST.get('top_heading')
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        TheFoundation.objects.create(top_heading=top_heading,heading=heading,image=image,description=description)
        return redirect('thefoundation_details')
    

class DetailsTheFoundation(View):
    def get(self,request):
        data = TheFoundation.objects.all()
        return render(request,'details_thefoundation.html',{'data':data})
    

class DeleteTheFoundation(View):
    def get(self,request,id):
        data = TheFoundation.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('thefoundation_details')
    

class EditTheFoundation(View):
    def get(self,request,id):
        data = TheFoundation.objects.get(id=id)
        return render(request,'edit_thefoundation.html',{'data':data})
    def post(self,request,id):
        data = TheFoundation.objects.get(id=id)
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        top_heading = request.POST.get('top_heading')
        if image is not None:
            TheFoundation(id=id,top_heading=top_heading,heading=heading,image=image,description=description).save()
        else:
            TheFoundation(id=id,top_heading=top_heading,heading=heading,image=data.image,description=description).save()
        return redirect('thefoundation_details')





################  overview  views  start here ################
class AddOverview(View):
    def get(self,request):
        data = Overview.objects.all()
        return render(request,'add_overview.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        description = request.POST.get('description')
        top_heading = request.POST.get('top_heading')
        top_description = request.POST.get('top_description')
        Overview.objects.create(top_heading=top_heading,heading=heading,top_description=top_description,description=description)
        return redirect('overview_details')
    

class DetailsOverview(View):
    def get(self,request):
        data = Overview.objects.all()
        return render(request,'details_overview.html',{'data':data})

class DeleteOverview(View):
    def get(self,request,id):
        data = Overview.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('overview_details')
    

class EditOverview(View):
    def get(self,request,id):
        data = Overview.objects.get(id=id)
        return render(request,'edit_overview.html',{'data':data})
    def post(self,request,id):
        data = Overview.objects.get(id=id)
        heading = request.POST.get('heading')
        description = request.POST.get('description')
        top_heading = request.POST.get('top_heading')
        top_description = request.POST.get('top_description')
        Overview(id=id,top_heading=top_heading,heading=heading,top_description=top_description,description=description).save()
        Overview(id=id,top_heading=top_heading,heading=heading,top_description=top_description,description=description).save()
        return redirect('overview_details')



################# The book views start here 
class AddTheBook(View):
    def get(self,request):
        data = AboutTheBook.objects.all()
        return render(request,'add_thebook.html',{'data':data})
    def post(self,request):
        title = request.POST.get('title')
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        AboutTheBook.objects.create(title=title,heading=heading,image=image,description=description)
        return redirect('thebook_details')
    

class DetailsTheBook(View):
    def get(self,request):
        data = AboutTheBook.objects.all()
        return render(request,'details_thebook.html',{'data':data})
    

class DeleteTheBook(View):
    def get(self,request,id):
        data = AboutTheBook.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('thebook_details')
    

class EditTheBook(View):
    def get(self,request,id):
        data = AboutTheBook.objects.get(id=id)
        return render(request,'edit_thefoundation.html',{'data':data})
    def post(self,request,id):
        data = AboutTheBook.objects.get(id=id)
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        title = request.POST.get('title')
        if image is not None:
            AboutTheBook(id=id,title=title,heading=heading,image=image,description=description).save()
        else:
            AboutTheBook(id=id,title=title,heading=heading,image=data.image,description=description).save()
        return redirect('thebook_details')





######### blogs views start here ..
class AddBlogBanner(View):
    def get(self,request):
        data = BlogsBanner.objects.all()
        return render(request,'add_blogsbanner.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        banner_image = request.FILES.get('banner_image')
        description = request.POST.get('description')
        BlogsBanner.objects.create(heading=heading,banner_image=banner_image,description=description)
        return redirect('BlogBanner_details')
    

class DetailsBlog(View):
    def get(self,request):
        data = BlogsBanner.objects.all()
        return render(request,'details_blogbanner.html',{'data':data})
    

class DeleteBlog(View):
    def get(self,request,id):
        data = BlogsBanner.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('BlogBanner_details')
    

class EditBlogBanner(View):
    def get(self,request,id):
        data = BlogsBanner.objects.get(id=id)
        return render(request,'edit_aboutbanner.html',{'data':data})
    def post(self,request,id):
        data = BlogsBanner.objects.get(id=id)
        heading = request.POST.get('heading')
        banner_image = request.FILES.get('banner_image')
        description = request.POST.get('description')
        if banner_image is not None:
            BlogsBanner(id=id,heading=heading,banner_image=banner_image,description=description).save()
        else:
            BlogsBanner(id=id,heading=heading,banner_image=data.banner_image,description=description).save()
        return redirect('BlogBanner_details')




############## blogs details views start here 


class AddBlogDetails(View):
    def get(self,request):
        data = BlogsDetails.objects.all()
        return render(request,'add_blogsdetail.html',{'data':data})
    def post(self,request):
        paragraph1=request.POST.get('paragraph1')
        paragraph2=request.POST.get('paragraph2')
        paragraph3=request.POST.get('paragraph3')
        title = request.POST.get('title')
        like = request.POST.get('like')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        by_name = request.POST.get('by_name')
        date = request.POST.get('date')
        month = request.POST.get('month')
        video = request.FILES.get('video')
        BlogsDetails.objects.create(video=video,date=date,month=month,by_name=by_name,title=title,paragraph1=paragraph1,paragraph2=paragraph2,paragraph3=paragraph3,like=like,image=image,description=description)
        return redirect('Blogdetail_details')
    

class Details_Blogs(View):
    def get(self,request):
        data = BlogsDetails.objects.all()
        return render(request,'detail_blogsdetails.html',{'data':data})
    

class DeleteBlogDetails(View):
    def get(self,request,id):
        data = BlogsDetails.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('Blogdetail_details')
    

class EditBlogDetail(View):
    def get(self,request,id):
        data = BlogsDetails.objects.get(id=id)
        return render(request,'edit_blogdetails.html',{'data':data})
    def post(self,request,id):
        data = BlogsDetails.objects.get(id=id)
        paragraph1=request.POST.get('paragraph1')
        paragraph2=request.POST.get('paragraph2')
        paragraph3=request.POST.get('paragraph3')
        title = request.POST.get('title')
        like = request.POST.get('like')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        by_name = request.POST.get('by_name')
        date = request.POST.get('date')
        month = request.POST.get('month')
        video = request.FILES.get('video')
        if video:
            data.video=video
            data.save()
        if image:
            data.image=image
            data.date=date
            data.video=video
            data.month=month
            data.by_name=by_name
            data.title=title
            data.paragraph1=paragraph1
            data.paragraph2=paragraph2
            data.paragraph3=paragraph3
            data.like=like
            data.image=image
            data.description=description
            data.save()
        return redirect('Blogdetail_details')




class RequestAdemoDetails(View):
    def get(self,request):
        data = RequestADemo.objects.all()
        return render(request,'requestAdemodetail.html',{'data':data})

class RequestDemoDelete(View):
    def get(self,request,id):
        data = RequestADemo.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('Demo_details')




###### digital publishing banner views start here..
class AddDigitalBanner(View):
    def get(self,request):
        data = DigitalPubBanner.objects.all()
        return render(request,'add_digitalpubbanner.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        banner_image = request.FILES.get('banner_image')
        description = request.POST.get('description')
        DigitalPubBanner.objects.create(heading=heading,banner_image=banner_image,description=description)
        return redirect('digital_details')
    

class DetailsDigital(View):
    def get(self,request):
        data = DigitalPubBanner.objects.all()
        return render(request,'detail_digitalpubbanner.html',{'data':data})
    

class DeleteDigital(View):
    def get(self,request,id):
        data = DigitalPubBanner.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('digital_details')
    

class EditDigitalBanner(View):
    def get(self,request,id):
        data = DigitalPubBanner.objects.get(id=id)
        return render(request,'edit_digitalpubbanner.html',{'data':data})
    def post(self,request,id):
        data = DigitalPubBanner.objects.get(id=id)
        heading = request.POST.get('heading')
        banner_image = request.FILES.get('banner_image')
        description = request.POST.get('description')
        if banner_image is not None:
            DigitalPubBanner(id=id,heading=heading,banner_image=banner_image,description=description).save()
        else:
            DigitalPubBanner(id=id,heading=heading,banner_image=data.banner_image,description=description).save()
        return redirect('digital_details')
    
######## Cater to Various #########
class AddCatertovarious(View):
    def get(self,request):
        data = CaterToVarious.objects.all()
        return render(request,'add_caterto.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        title = request.POST.get('title')
        span = request.POST.get('span')
        CaterToVarious.objects.create(title=title,span=span,heading=heading,image=image,description=description)
        return redirect('Caterto_details')
    

class DetailsCaterto(View):
    def get(self,request):
        data = CaterToVarious.objects.all()
        return render(request,'detail_caterto.html',{'data':data})
    

class DeleteCaterto(View):
    def get(self,request,id):
        data = CaterToVarious.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('Caterto_details')
    

class EditCaterto(View):
    def get(self,request,id):
        data = CaterToVarious.objects.get(id=id)
        return render(request,'edit_caterto.html',{'data':data})
    def post(self,request,id):
        data = CaterToVarious.objects.get(id=id)
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        title = request.POST.get('title')
        span = request.POST.get('span')
        if image is not None:
            CaterToVarious(id=id,title=title,span=span,heading=heading,image=image,description=description).save()
        else:
            CaterToVarious(id=id,title=title,span=span,heading=heading,image=data.image,description=description).save()
        return redirect('Caterto_details')





########## Secured distribution views start here..
class AddSecured(View):
    def get(self,request):
        data = SecuredDistribution.objects.all()
        return render(request,'add_secured.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        title = request.POST.get('title')
        span = request.POST.get('span')
        SecuredDistribution.objects.create(title=title,span=span,heading=heading,image=image,description=description)
        return redirect('Secured_details')
    
class DetailsSecured(View):
    def get(self,request):
        data = SecuredDistribution.objects.all()
        return render(request,'detail_secured.html',{'data':data})
    
class DeleteSecured(View):
    def get(self,request,id):
        data = SecuredDistribution.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('Secured_details')
    


class EditSecured(View):
    def get(self,request,id):
        data = SecuredDistribution.objects.get(id=id)
        return render(request,'edit_secured.html',{'data':data})
    def post(self,request,id):
        data = SecuredDistribution.objects.get(id=id)
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        title = request.POST.get('title')
        span = request.POST.get('span')
        if image is not None:
            SecuredDistribution(id=id,title=title,span=span,heading=heading,image=image,description=description).save()
        else:
            SecuredDistribution(id=id,title=title,span=span,heading=heading,image=data.image,description=description).save()
        return redirect('Secured_details')



########### get your own branded 
class AddGetYour(View):
    def get(self,request):
        data = GetYourBranded.objects.all()
        return render(request,'add_getyour.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        title = request.POST.get('title')
        span = request.POST.get('span')
        GetYourBranded.objects.create(title=title,span=span,heading=heading,image=image,description=description)
        return redirect('GetYour_details')
    
class DetailsGetYour(View):
    def get(self,request):
        data = GetYourBranded.objects.all()
        return render(request,'detail_getyour.html',{'data':data})
    
class DeleteGetYour(View):
    def get(self,request,id):
        data = GetYourBranded.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('GetYour_details')
    


class EditGetYour(View):
    def get(self,request,id):
        data = GetYourBranded.objects.get(id=id)
        return render(request,'edit_getyour.html',{'data':data})
    def post(self,request,id):
        data = GetYourBranded.objects.get(id=id)
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        title = request.POST.get('title')
        span = request.POST.get('span')
        if image is not None:
            GetYourBranded(id=id,title=title,span=span,heading=heading,image=image,description=description).save()
        else:
            GetYourBranded(id=id,title=title,span=span,heading=heading,image=data.image,description=description).save()
        return redirect('GetYour_details')


######## Digital Publish on Multi-Devices views start here..
class AddMultiDevice(View):
    def get(self,request):
        data = MultiDevices.objects.all()
        return render(request,'add_multidevice.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        small_heading = request.POST.get('small_heading')
        icon = request.POST.get('icon')
        small_description = request.POST.get('small_description')
        MultiDevices.objects.create(small_description=small_description,small_heading=small_heading,icon=icon,heading=heading,image=image,description=description)
        return redirect('MultiDevice_details')
    
class DetailsMultiDevice(View):
    def get(self,request):
        data = MultiDevices.objects.all()
        return render(request,'details_multidevice.html',{'data':data})
    
class DeleteMultiDevice(View):
    def get(self,request,id):
        data = MultiDevices.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('MultiDevice_details')

class EditMultiDevice(View):
    def get(self,request,id):
        data = MultiDevices.objects.get(id=id)
        return render(request,'edit_multidevice.html',{'data':data})
    def post(self,request,id):
        data = MultiDevices.objects.get(id=id)
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        title = request.POST.get('title')
        icon = request.POST.get('icon')
        small_description = request.POST.get('small_description')
        if image is not None:
            MultiDevices(id=id,small_description=small_description,title=title,icon=icon,heading=heading,image=image,description=description).save()
        else:
            MultiDevices(id=id,small_description=small_description,title=title,icon=icon,heading=heading,image=data.image,description=description).save()
        return redirect('MultiDevice_details')


############# Easily integrates views start here 
class AddEasilyInter(View):
    def get(self,request):
        data = Easilyintegrates.objects.all()
        return render(request,'add_easilyinter.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        Easilyintegrates.objects.create(heading=heading,image=image,description=description)
        return redirect('EasilyInter_details')
    

class DetailsEasilyInter(View):
    def get(self,request):
        data = Easilyintegrates.objects.all()
        return render(request,'details_easilyinter.html',{'data':data})
    

class DeleteEasilyInter(View):
    def get(self,request,id):
        data = Easilyintegrates.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('EasilyInter_details')
    

class EditEasilyInter(View):
    def get(self,request,id):
        data = Easilyintegrates.objects.get(id=id)
        return render(request,'edit_digitalpubbanner.html',{'data':data})
    def post(self,request,id):
        data = Easilyintegrates.objects.get(id=id)
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        if image is not None:
            Easilyintegrates(id=id,heading=heading,image=image,description=description).save()
        else:
            Easilyintegrates(id=id,heading=heading,image=data.image,description=description).save()
        return redirect('EasilyInter_details')
    

############### e-book store views 
class AddBookStoreBanner(View):
    def get(self,request):
        data = BookStoreBanner.objects.all()
        return render(request,'add-bookstorebanner.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        banner_image = request.FILES.get('banner_image')
        description = request.POST.get('description')
        BookStoreBanner.objects.create(heading=heading,banner_image=banner_image,description=description)
        return redirect('bookstore_details')
    

class DetailsBookStore(View):
    def get(self,request):
        data = BookStoreBanner.objects.all()
        return render(request,'details_bookstorebanner.html',{'data':data})
    

class DeleteBookStore(View):
    def get(self,request,id):
        data = BookStoreBanner.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('bookstore_details')
    

class EditBookStoreBanner(View):
    def get(self,request,id):
        data = BookStoreBanner.objects.get(id=id)
        return render(request,'edit_bookstorebanner.html',{'data':data})
    def post(self,request,id):
        data = BookStoreBanner.objects.get(id=id)
        heading = request.POST.get('heading')
        banner_image = request.FILES.get('banner_image')
        description = request.POST.get('description')
        if banner_image is not None:
            BookStoreBanner(id=id,heading=heading,banner_image=banner_image,description=description).save()
        else:
            BookStoreBanner(id=id,heading=heading,banner_image=data.banner_image,description=description).save()
        return redirect('bookstore_details')




##### Your branded webstore views 
class AddYourWebstore(View):
    def get(self,request):
        data = YourBrandedWebstore.objects.all()
        return render(request,'add_yourwebstore.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        title = request.POST.get('title')
        span = request.POST.get('span')
        YourBrandedWebstore.objects.create(title=title,span=span,heading=heading,image=image,description=description)
        return redirect('YourWebstore_details')

class DetailsYourWebstore(View):
    def get(self,request):
        data = YourBrandedWebstore.objects.all()
        return render(request,'details_yourwebstore.html',{'data':data})
    
class DeleteYourWebstore(View):
    def get(self,request,id):
        data = YourBrandedWebstore.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('YourWebstore_details')
    


class EditYourWebstore(View):
    def get(self,request,id):
        data = YourBrandedWebstore.objects.get(id=id)
        return render(request,'edit_yourwebstore.html',{'data':data})
    def post(self,request,id):
        data = YourBrandedWebstore.objects.get(id=id)
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        title = request.POST.get('title')
        span = request.POST.get('span')
        if image is not None:
            YourBrandedWebstore(id=id,title=title,span=span,heading=heading,image=image,description=description).save()
        else:
            YourBrandedWebstore(id=id,title=title,span=span,heading=heading,image=data.image,description=description).save()
        return redirect('YourWebstore_details')



######## Easy Integration  views start here ########
class AddEasyIntegration(View):
    def get(self,request):
        data = EasyIntegration.objects.all()
        return render(request,'add_easyintegration.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        title = request.POST.get('title')
        span = request.POST.get('span')
        EasyIntegration.objects.create(title=title,span=span,heading=heading,image=image,description=description)
        return redirect('EasyIntegration_details')


class DetailsEasyIntegration(View):
    def get(self,request):
        data = EasyIntegration.objects.all()
        return render(request,'detail_easyintegrations.html',{'data':data})
    
class DeleteEasyIntegration(View):
    def get(self,request,id):
        data = EasyIntegration.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('EasyIntegration_details')
    

class EditEasyIntegration(View):
    def get(self,request,id):
        data = EasyIntegration.objects.get(id=id)
        return render(request,'edit_easyintegrations.html',{'data':data})
    def post(self,request,id):
        data = EasyIntegration.objects.get(id=id)
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        title = request.POST.get('title')
        span = request.POST.get('span')
        if image is not None:
            EasyIntegration(id=id,title=title,span=span,heading=heading,image=image,description=description).save()
        else:
            EasyIntegration(id=id,title=title,span=span,heading=heading,image=data.image,description=description).save()
        return redirect('EasyIntegration_details')




######### social network view start here ######
class AddSocialNetwork(View):
    def get(self,request):
        data = SocialNetwork.objects.all()
        return render(request,'add_socialnetwork.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description1 = request.POST.get('description1')
        title = request.POST.get('title')
        span = request.POST.get('span')
        description2 = request.POST.get('description2')
        SocialNetwork.objects.create(title=title,span=span,heading=heading,image=image,description1=description1,description2=description2)
        return redirect('SocialNetwork_details')

class DetailsSocialNetwork(View):
    def get(self,request):
        data = SocialNetwork.objects.all()
        return render(request,'details_socialnetwork.html',{'data':data})
    
class DeleteSocialNetwork(View):
    def get(self,request,id):
        data = SocialNetwork.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('SocialNetwork_details')
    
class EditSocialNetwork(View):
    def get(self,request,id):
        data = SocialNetwork.objects.get(id=id)
        return render(request,'edit_socialnetwork.html',{'data':data})
    def post(self,request,id):
        data = SocialNetwork.objects.get(id=id)
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description1 = request.POST.get('description1')
        title = request.POST.get('title')
        span = request.POST.get('span')
        description2 = request.POST.get('description2')
        if image is not None:
            SocialNetwork(id=id,title=title,span=span,heading=heading,image=image,description1=description1,description2=description2).save()
        else:
            SocialNetwork(id=id,title=title,span=span,heading=heading,image=data.image,description1=description1,description2=description2).save()
        return redirect('SocialNetwork_details')


####### fully responsive views start here..
class AddFullyResponsive(View):
    def get(self,request):
        data = FullyResponsive.objects.all()
        return render(request,'add_fullyresponsive.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        FullyResponsive.objects.create(heading=heading,image=image,description=description)
        return redirect('FullyResponsive_details')
    

class DetailsFullyResponsive(View):
    def get(self,request):
        data = FullyResponsive.objects.all()
        return render(request,'details_fullyresponsive.html',{'data':data})
    

class DeleteFullyResponsive(View):
    def get(self,request,id):
        data = FullyResponsive.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('FullyResponsive_details')


class EditFullyResponsive(View):
    def get(self,request,id):
        data = FullyResponsive.objects.get(id=id)
        return render(request,'edit_fullyresponsive.html',{'data':data})
    def post(self,request,id):
        data = FullyResponsive.objects.get(id=id)
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        if image is not None:
            FullyResponsive(id=id,heading=heading,image=image,description=description).save()
        else:
            FullyResponsive(id=id,heading=heading,image=data.image,description=description).save()
        return redirect('FullyResponsive_details')






####### terms and conditions views start here 
class AddTermCondition(View):
    def get(self,request):
        data = TermsCondition.objects.all()
        return render(request,'add_term&condition.html',{'data':data})
    def post(self,request):
        image = request.FILES.get('image')
        description = request.POST.get('description')
        TermsCondition.objects.create(description=description,image=image)
        return redirect('TermCondition_details')

class DetailsTermCondition(View):
    def get(self,request):
        data = TermsCondition.objects.all()
        return render(request,'details_term&condition.html',{'data':data})
    

class EditTermCondition(View):
    def get(self,request,id):
        data = TermsCondition.objects.get(id=id)
        return render(request,'edit_term&condition.html',{'data':data})
    def post(self,request,id):
        data = TermsCondition.objects.get(id=id)
        description = request.POST.get('description')
        image = request.FILES.get('image')
        if image:
            TermsCondition(id=id,description=description,image=image).save()
        else:
            TermsCondition(id=id,description=description,image=data.image).save()
        return redirect('TermCondition_details')


class DeleteTermCondition(View):
    def get(self,request,id):
        data = TermsCondition.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully')
        return redirect('TermCondition_details')


###### privacy &  policy views start here..
class AddPrivacy(View):
    def get(self,request):
        data = PrivacyPolicy.objects.all()
        return render(request,'add_privacypolicy.html',{'data':data})
    def post(self,request):
        image = request.FILES.get('image')
        description = request.POST.get('description')
        PrivacyPolicy.objects.create(description=description,image=image)
        return redirect('Privacy_details')

class DetailsPrivacy(View):
    def get(self,request):
        data = PrivacyPolicy.objects.all()
        return render(request,'details_privacypolicy.html',{'data':data})
    

class EditPrivacy(View):
    def get(self,request,id):
        data = PrivacyPolicy.objects.get(id=id)
        return render(request,'edit_privacypolicy.html',{'data':data})
    def post(self,request,id):
        data = PrivacyPolicy.objects.get(id=id)
        description = request.POST.get('description')
        image = request.FILES.get('image')
        if image:
            PrivacyPolicy(id=id,description=description,image=image).save()
        else:
            PrivacyPolicy(id=id,description=description,image=data.image).save()
        return redirect('Privacy_details')


class DeletePrivacy(View):
    def get(self,request,id):
        data = PrivacyPolicy.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully')
        return redirect('Privacy_details')












##### faq views write here..
class AddFaq(View):
    def get(self,request):
        data = Faq.objects.all()
        return render(request,'add_faq.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        description = request.POST.get('description')
        question = request.POST.get('question')
        ans = request.POST.get('ans')
        image = request.FILES.get('image')
        Faq.objects.create(heading=heading,description=description,question=question,ans=ans,image=image)
        return redirect('Faq_details')


class DetailsFaq(View):
    def get(self,request):
        data = Faq.objects.all()
        return render(request,'details_faq.html',{'data':data})


class EditFaq(View):
    def get(self,request,id):
        data = Faq.objects.get(id=id)
        return render(request,'edit_faq.html',{'data':data})
    def post(self,request,id):
        data = Faq.objects.get(id=id)
        heading = request.POST.get('heading')
        description = request.POST.get('description')
        question = request.POST.get('question')
        ans = request.POST.get('ans')
        image = request.FILES.get('image')
        if image is not None:
            Faq(id=id,heading=heading,description=description,question=question,ans=ans,image=image).save()
        else:
            Faq(id=id,heading=heading,description=description,question=question,ans=ans,image=data.image).save()
        return redirect('Faq_details')


class DeleteFaq(View):
    def get(self,request,id):
        data = Faq.objects.get(id=id)
        data.delete('')
        messages.success(request,'Data Deleted Successfully')
        return redirect('Faq_details')



###### add notification views start here..
class AddNotification(View):
    def get(self, request):
        data = Notifications.objects.all()
        users = User.objects.filter(Q(is_student=True) | Q(is_organization=True))
        students = users.filter(is_student=True)
        corporates = users.filter(is_organization=True)
        return render(request, 'add_notification.html', {'data': data, 'students': students, 'corporates': corporates})
    def post(self, request):
        receiver_emails = request.POST.getlist('receiver[]')
        subject = request.POST.get('subject')
        message = request.POST.get('message')

        for email in receiver_emails:
            if email == 'all_students':
                # Send to all students
                students = User.objects.filter(is_student=True)
                for student in students:
                    Notifications.objects.create(receiver=student, subject=subject, message=message)
            elif email == 'all_corporates':
                # Send to all corporates
                corporates = User.objects.filter(is_organization=True)
                for corporate in corporates:
                    Notifications.objects.create(receiver=corporate, subject=subject, message=message)
            else:
                # Send to individual recipients
                try:
                    user = User.objects.get(email=email.strip())
                    Notifications.objects.create(receiver=user, subject=subject, message=message)
                except User.DoesNotExist:
                    # Handle the case where the user does not exist
                    pass

        return redirect('detailsnotification')




class DetailNotification(View):
    def get(self,request):
        data = Notifications.objects.all()
        return render(request,'details_notification.html',{'data':data,'active28':'active12'})
    

class DeleteNotification(View):
    def get(self,request,id):
        data = Notifications.objects.get(id=id)
        data.delete()
        return redirect('detailsnotification')





############ ebook reader views start here.
class Addebookreader(View):
    def get(self,request):
        data = EbookReaderBanner.objects.all()
        return render(request,'add_ebookreader.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        EbookReaderBanner.objects.create(heading=heading,image=image,description=description)
        return redirect('ebookreader_details')
    

class Detailsebookreader(View):
    def get(self,request):
        data = EbookReaderBanner.objects.all()
        return render(request,'details_ebookreader.html',{'data':data})
    


class Deleteebookreader(View):
    def get(self,request,id):
        data = EbookReaderBanner.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('ebookreader_details')
    
class Editebookreader(View):
    def get(self,request,id):
        data = EbookReaderBanner.objects.get(id=id)
        return render(request,'edit_ebookreader.html',{'data':data})
    def post(self,request,id):
        data = EbookReaderBanner.objects.get(id=id)
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        if image is not None:
            EbookReaderBanner(id=id,heading=heading,image=image,description=description).save()
        else:
            EbookReaderBanner(id=id,heading=heading,image=data.image,description=description).save()
        return redirect('ebookreader_details')


######## indybot views start here 
class AddIndybotBanner(View):
    def get(self,request):
        data = IndyBotBanner.objects.all()
        return render(request,'add_indybotbanner.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        IndyBotBanner.objects.create(heading=heading,image=image,description=description)
        return redirect('indybot_details')
    

class DetailsIndybotBanner(View):
    def get(self,request):
        data = IndyBotBanner.objects.all()
        return render(request,'details_indybotbanner.html',{'data':data})
    

class DeleteIndybotBanner(View):
    def get(self,request,id):
        data = IndyBotBanner.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('indybot_details')
    

class EditIndybotBanner(View):
    def get(self,request,id):
        data = IndyBotBanner.objects.get(id=id)
        return render(request,'edit_indybot.html',{'data':data})
    def post(self,request,id):
        data = IndyBotBanner.objects.get(id=id)
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        if image is not None:
            IndyBotBanner(id=id,heading=heading,image=image,description=description).save()
        else:
            IndyBotBanner(id=id,heading=heading,image=data.image,description=description).save()
        return redirect('indybot_details')


############# #######################3
class AddUseprimary(View):
    def get(self,request): 
        data = Useprimary.objects.all()
        return render(request,'add_useprimary.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        image1 = request.FILES.get('image1')
        image2 = request.FILES.get('image2')
        image3 = request.FILES.get('image3')
        paragraph1 = request.POST.get('paragraph1')
        paragraph2 = request.POST.get('paragraph2')
        Useprimary.objects.create(heading=heading,image3=image3,image1=image1,image2=image2,paragraph1=paragraph1,paragraph2=paragraph2)
        return redirect('useprimary_details')
    

class DetailsUseprimary(View):
    def get(self,request):
        data = Useprimary.objects.all()
        return render(request,'detail_useprimary.html',{'data':data})
    

class DeleteUseprimary(View):
    def get(self,request,id):
        data = Useprimary.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('useprimary_details')
    

class EditUseprimary(View):
    def get(self, request, id):
        data = Useprimary.objects.get(id=id)
        return render(request, 'edit_useprimary.html', {'data': data})

    def post(self, request, id):
        data = Useprimary.objects.get(id=id)
        heading = request.POST.get('heading')
        image1 = request.FILES.get('image1')
        image2 = request.FILES.get('image2')
        image3 = request.FILES.get('image3')
        paragraph1 = request.POST.get('paragraph1')
        paragraph2 = request.POST.get('paragraph2')
        data.heading = heading
        data.paragraph1 = paragraph1
        data.paragraph2 = paragraph2
        if image1 is not None:
            data.image1 = image1
        if image2 is not None:
            data.image2 = image2
        if image3 is not None:
            data.image3 = image3

        data.save()
        return redirect('useprimary_details')



########## Indybot Responsive 
class AddIndybotRespon(View):
    def get(self,request):
        data = IndybotResponsive.objects.all()
        return render(request,'add_indyrespon.html',{'data':data})
    def post(self,request):
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        IndybotResponsive.objects.create(heading=heading,image=image,description=description)
        return redirect('indyrespo_details')
    


class DetailsIndybotRespon(View):
    def get(self,request):
        data = IndybotResponsive.objects.all()
        return render(request,'detail-indybotresponsive.html',{'data':data})
    

class DeleteIndybotRespon(View):
    def get(self,request,id):
        data = IndybotResponsive.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('indyrespo_details')
    

class EditIndybotRespon(View):
    def get(self,request,id):
        data = IndybotResponsive.objects.get(id=id)
        return render(request,'edit_indyresponsive.html',{'data':data})
    def post(self,request,id):
        data = IndybotResponsive.objects.get(id=id)
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        if image is not None:
            IndybotResponsive(id=id,heading=heading,image=image,description=description).save()
        else:
            IndybotResponsive(id=id,heading=heading,image=data.image,description=description).save()
        return redirect('indyrespo_details')



################ readyour book 
class AddReadBook(View):
    def get(self,request):
        data = ReadYourBook.objects.all()
        return render(request,'add_readbook.html',{'data':data})
    def post(self,request):
        top_heading = request.POST.get('top_heading')
        title = request.POST.get('title')
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        ReadYourBook.objects.create(top_heading=top_heading,heading=heading,image=image,description=description,title=title)
        return redirect('Readbook_details')
    
class DetailsReadBook(View):
    def get(self,request):
        data = ReadYourBook.objects.all()
        return render(request,'details_readyourbook.html',{'data':data})
    

class DeleteReadBook(View):
    def get(self,request,id):
        data = ReadYourBook.objects.get(id=id)
        data.delete()
        messages.success(request,'Data Deleted Successfully!..')
        return redirect('Readbook_details')
    

class EditReadBook(View):
    def get(self,request,id):
        data = ReadYourBook.objects.get(id=id)
        return render(request,'edit_readyourbook.html',{'data':data})
    def post(self,request,id):
        data = ReadYourBook.objects.get(id=id)
        heading = request.POST.get('heading')
        image = request.FILES.get('image')
        description = request.POST.get('description')
        top_heading = request.POST.get('top_heading')
        title = request.POST.get('title')
        if image is not None:
            ReadYourBook(id=id,top_heading=top_heading,title=title,heading=heading,image=image,description=description).save()
        else:
            ReadYourBook(id=id,top_heading=top_heading,title=title,heading=heading,image=data.image,description=description).save()
        return redirect('Readbook_details')






########### is_staff section start here ######
def staff_login(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        user = authenticate(request=request, email=email, password=password)
        if user is not None:
            if User.objects.filter(email=email,is_staff=True).exists():
                dj_login(request, user)
                return redirect('dashboard')
            else:
                messages.warning(request,'You Are Not Staff')
        else:
            messages.warning(request, 'Invalid Email or Password')
    return render(request,'staff_login.html')

class AddStaff(View):
    def get(self, request):
        if request.user.is_authenticated:
            is_superuser = request.user.is_superuser
            if is_superuser == True:
                context = {
                    'is_superuser': is_superuser
                }
                return render(request, 'add_staff.html',{**context})
        return redirect('login')
    def post(self, request):
        User = get_user_model()
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        phone_number = request.POST.get('phone_number')
        password = request.POST.get('password')
        institute = request.POST.get('institute')
        if User.objects.filter(email=email,).exists():
            admin_message = f"A user email address '{email}' already exists. Please use another  email id  to create account."
            
            messages.error(request, admin_message)
            return redirect('addstaff')
        elif User.objects.filter(phone_number=phone_number,).exists():
            messages.error(request,f'Phone number  already exists {phone_number}')
            return redirect('addstaff')
        user = User.objects.create_user(institute=institute,first_name=first_name, last_name=last_name, email=email, phone_number=phone_number, password=password)
        user.is_staff = True
        user.is_active = True
        user.save()
        selected_permissions = request.POST.getlist('permissions_checkbox')
        staff_menu = Staffmenu()
        staff_menu.user = user

        staff_menu.dashboard = 'dashboard' in selected_permissions and '1' or '0'
        staff_menu.details_book = 'details_book' in selected_permissions and '1' or '0'
        staff_menu.user_details = 'user_details' in selected_permissions and '1' or '0'
        staff_menu.manage_home = 'manage_home' in selected_permissions and '1' or '0'
        staff_menu.manage_contact = 'manage_contact' in selected_permissions and '1' or '0'
        staff_menu.manage_category = 'manage_category' in selected_permissions and '1' or '0'
        staff_menu.manage_collection = 'manage_collection' in selected_permissions and '1' or '0'
        staff_menu.digital_publishing = 'digital_publishing' in selected_permissions and '1' or '0'
        staff_menu.ebook_store = 'ebook_store' in selected_permissions and '1' or '0'
        staff_menu.ebook_reader = 'ebook_reader' in selected_permissions and '1' or '0'
        staff_menu.indybot = 'indybot' in selected_permissions and '1' or '0'
        staff_menu.notifications = 'notifications' in selected_permissions and '1' or '0'
        staff_menu.manage_cms = 'manage_cms' in selected_permissions and '1' or '0'
        staff_menu.manage_staff = 'manage_staff' in selected_permissions and '1' or '0'
        staff_menu.manage_blogs = 'manage_blogs' in selected_permissions and '1' or '0'
        staff_menu.user = user
        staff_menu.save()
        return redirect('details_staff')




class BookDetails(View):
    def get(self, request, id):
        if request.user.is_authenticated:
            user = request.user.institute
            footer = Footer.objects.all()
            member = User.objects.filter(is_student=True, is_active=True, is_superuser=False, is_staff=False, is_organization=False).count()
            book_list = Books.objects.filter(id=id,is_active=True)  
            book_obj = Books.objects.get(id=id,is_active=True)  
            book = Books.objects.get(id=id,is_active=True)

            book.viewed_at = timezone.now()
            book.save() 
            book = Books.objects.get(id=id,is_active=True)
            add_recent_view_book = Recently_Viewed.objects.create(user=request.user, book=book)
            completed = False
            if BookVisit.objects.filter(user=request.user, books=book).exists():
                last_visit_book = BookVisit.objects.filter(user=request.user, books=book).last()
                
                if last_visit_book.completed == True:
                    completed = True
                else:
                    completed = False
            return render(request, 'organizations/book_deatils.html', { 'book_obj':book_obj, 'footer': footer,'book_list': book_list,'member':member, "id":id})
        else:
            messages.error(request, 'Login is required to access book, Please login if register account or register as new user')
            return redirect('home')
        
class DetailsStaff(View):
    def get(self,request):
        if request.user.is_authenticated :
            is_superuser = request.user.is_superuser 
            if is_superuser == True:
                context={
                    'is_superuser':is_superuser
                }
                data=User.objects.filter(is_staff=True).order_by('-id').exclude(is_superuser = True)
                return render(request,'details_staff.html',{'data':data, 'active27':'active12',**context})
        return redirect('login')


########### edit staff views ############
class edit_staff(View):
    def get(self,request,id):
        if request.user.is_authenticated : 
            is_superuser= request.user.is_superuser
            data =  User.objects.get(id=id)
            staff_menu = Staffmenu.objects.get(user=data.id)   
            if is_superuser == True:
                context= {
                    'is_superuser':is_superuser,
                    'data':data, 
                }  
                return render(request,'edit_staff.html',{**context,'staff_menu':staff_menu})
        return redirect('login')
    def post(self, request, id):
        user = User.objects.get(id=id)
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        phone_number = request.POST.get('phone_number')
        password = request.POST.get('password')
        selected_permissions = request.POST.getlist('permissions_checkbox')
        user.first_name = first_name
        user.last_name = last_name
        user.email = email
        user.phone_number = phone_number

        if password:
            user.password = make_password(password)  # Hash and save the password
        user.is_staff = True
        user.save()
        # Update staff_menu
        data = User.objects.get(id=id)
        staff_menu = Staffmenu.objects.get(user=data.id)
        staff_menu.dashboard = 'dashboard' in selected_permissions and '1' or '0'
        staff_menu.details_book = 'details_book' in selected_permissions and '1' or '0'
        staff_menu.user_details = 'user_details' in selected_permissions and '1' or '0'
        staff_menu.manage_home = 'manage_home' in selected_permissions and '1' or '0'
        staff_menu.manage_contact = 'manage_contact' in selected_permissions and '1' or '0'
        staff_menu.manage_category = 'manage_category' in selected_permissions and '1' or '0'
        staff_menu.manage_collection = 'manage_collection' in selected_permissions and '1' or '0'
        staff_menu.digital_publishing = 'digital_publishing' in selected_permissions and '1' or '0'
        staff_menu.ebook_store = 'ebook_store' in selected_permissions and '1' or '0'
        staff_menu.ebook_reader = 'ebook_reader' in selected_permissions and '1' or '0'
        staff_menu.indybot = 'indybot' in selected_permissions and '1' or '0'
        staff_menu.notifications = 'notifications' in selected_permissions and '1' or '0'
        staff_menu.manage_cms = 'manage_cms' in selected_permissions and '1' or '0'
        staff_menu.manage_staff = 'manage_staff' in selected_permissions and '1' or '0'
        staff_menu.manage_blogs = 'manage_blogs' in selected_permissions and '1' or '0'
        staff_menu.save()
        return redirect('details_staff')




####### staff active inactive views 
class Active_inActiveStaff(View):
    def post(self, request):
        user_id = request.POST['id']
        user = User.objects.get(id=user_id)
        if user.is_active is False:
            user.is_active = True
            user.save()
            return redirect('details_staff')
        elif user.is_active is True:
            user.is_active = False
            user.save()
            return redirect('details_staff')
        else:
            return HttpResponse("User Not Valid")




########### corporate views start here........
class CreateCorporate(View):
    def get(self, request):
        data = User.objects.all()
        books = Books.objects.all()
        return render(request, 'organizations/create_corporate.html', {'data': data, 'books': books})

    def post(self, request):
        institute = request.POST.get('institute')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        phone_number = request.POST.get('phone_number')
        password = request.POST.get('password')
        selected_books = request.POST.getlist('books')

        if User.objects.filter(email=email).exists():
            messages.error(request, 'Email id Already Exists. Please use a different email to create an account.')
            return redirect('Create_organizations')  # Replace 'your_redirect_url' with the appropriate URL
        if User.objects.filter(phone_number=phone_number).exists():
            messages.error(request, 'Phone number Already Exists. Please use a different phone number to create an account.')
            return redirect('Create_organizations')
        user = User.objects.create_user(
            institute=institute,
            first_name=first_name,
            last_name=last_name,
            email=email,
            phone_number=phone_number,
            password=password
        )
        user.is_organization = True
        user.is_active = True
        user.save()

        subject = 'Your Corporate Account Details'
        message = f'Hello {user.first_name},\n\nYour corporate account has been created.\n\nEmail: {user.email}\nPassword: {password}\n\nYou can access your account at [http://64.227.157.173:8002/Admin/].\n\nThank you!'
        from_email = 'indy@optimuse-solutions.com'
        recipient_list = [user.email]
        send_mail(subject, message, from_email, recipient_list, fail_silently=False)

        for book_id in selected_books:
            selected_book = Books.objects.get(id=book_id)
            selected_book.is_organization = True
            selected_book.save()
            user.books.add(selected_book)  # Use add() to add the selected book to the ManyToManyField

        messages.success(request, 'Organization Created Successfully.')
        return redirect('Detail_organizations')  # Replace 'Detail_organizations' with the appropriate URL


class DeleteStaff(View):
    def get(self,request,id):
        data = User.objects.get(id=id)
        data.delete()
        return redirect('details_staff')

class MostPopularCorporateBookList(View):
    def get(self, request):
        institute_name = getattr(request.user, 'institute', 'guest')  # Use getattr to provide a default value
        book_visits = BookVisit.objects.filter(institute_name=institute_name)

        total_time_per_book = defaultdict(int)
        visits_per_book = defaultdict(int)
        num_books_to_process = 5  # Set the number of books you want to process

        for visit in book_visits:
            user_visit = BookVisit.objects.filter(books=visit.books, institute_name=request.user.institute).count()
            visits_per_book[visit.books_id] += 1  # Increment the visit count for the book

            if visit.start_time and visit.end_time:
                time_difference = visit.end_time - visit.start_time
                total_time_per_book[visit.books_id] += time_difference.total_seconds()

        # Sort books based on total time spent in descending order
        sorted_data_desc = sorted(total_time_per_book.items(), key=lambda item: item[1], reverse=True)[:num_books_to_process]
        
        # Create a dictionary to store formatted times, visit count, and book objects for each book
        formatted_data_per_book = {}
        for book_id, total_time in sorted_data_desc:
            hours, remainder = divmod(total_time, 3600)
            minutes = remainder / 60
            if hours >= 1:
                formatted_time = "{:.0f} h {:.0f} min".format(hours, minutes)
            else:
                formatted_time = "{:.0f} min".format(minutes)

            try:
                book_object = Books.objects.get(id=book_id)
            except Books.DoesNotExist:
                book_object = None

            formatted_data_per_book[book_id] = {
                'formatted_time': formatted_time,
                'visit_count': visits_per_book[book_id],  # Add visit count to the dictionary
                'book_object': book_object,
               
            }

        context = {'formatted_data_per_book': formatted_data_per_book, 'active14':'active20'}
        return render(request, 'organizations/most-popular-books.html', context)


class LeastPopularCorporateBookList(View):
    def get(self, request):
        institute_name = getattr(request.user, 'institute', 'guest')  # Use getattr to provide a default value
        book_visits = BookVisit.objects.filter(institute_name=institute_name)

        total_time_per_book = defaultdict(int)
        visits_per_book = defaultdict(int)
        num_books_to_process = 5  # Set the number of books you want to process

        for visit in book_visits:
            user_visit = BookVisit.objects.filter(books=visit.books, institute_name=request.user.institute).count()
            visits_per_book[visit.books_id] += 1  # Increment the visit count for the book

            if visit.start_time and visit.end_time:
                time_difference = visit.end_time - visit.start_time
                total_time_per_book[visit.books_id] += time_difference.total_seconds()

        # Sort books based on total time spent in descending order
        sorted_data_desc = sorted(total_time_per_book.items(), key=lambda item: item[1], reverse=False)[:num_books_to_process]
        
        # Create a dictionary to store formatted times, visit count, and book objects for each book
        formatted_data_per_book = {}
        for book_id, total_time in sorted_data_desc:
            hours, remainder = divmod(total_time, 3600)
            minutes = remainder / 60
            if hours >= 1:
                formatted_time = "{:.0f} h {:.0f} min".format(hours, minutes)
            else:
                formatted_time = "{:.0f} min".format(minutes)

            try:
                book_object = Books.objects.get(id=book_id)
            except Books.DoesNotExist:
                book_object = None

            formatted_data_per_book[book_id] = {
                'formatted_time': formatted_time,
                'visit_count': visits_per_book[book_id],  # Add visit count to the dictionary
                'book_object': book_object,
            }

        context = {'formatted_data_per_book': formatted_data_per_book}
        return render(request,'organizations/least-poplular-books.html',context)

# class AverageTimeReadingBooks(View):
#     def get(self, request):
#         # Get user's books and institute name
#         user_books = request.user.books.all()
#         institute_name = request.user.institute  # Default to 'guest' if not provided

#         # Fetch book visits for the given institute
#         book_visits = BookVisit.objects.filter(institute_name=institute_name)

#         # Use defaultdict to store total time for each book
#         total_time_per_book = defaultdict(int)

#         # Set the number of books you want to process
#         num_books_to_process = 10  

#         # Process book visits
#         for visit in book_visits:
#             if visit.start_time and visit.end_time:
#                 time_difference = visit.end_time - visit.start_time
#                 total_time_per_book[visit.books_id] += time_difference.total_seconds()

#             if len(total_time_per_book) >= num_books_to_process:
#                 break

#         # Create a dictionary to store formatted times and book objects for each book
#         formatted_data_per_book = {}

#         # Convert total_time to hours and minutes for each book
#         for book_id, total_time in total_time_per_book.items():
#             try:
#                 book_object = Books.objects.get(id=book_id)

#                 # Calculate the unique count of users who visited the book for the given institute
#                 book_visits_count = BookVisit.objects.filter(books=book_object, institute_name=institute_name).count()

#             except Books.DoesNotExist:
#                 # Handle the case where the book is not found
#                 book_object = None

#             # Avoid division by zero
#             if book_visits_count > 0:
#                 time_per_user = total_time / book_visits_count
#             else:
#                 time_per_user = 0

#             # Convert the result into hours and minutes
#             hours, remainder = divmod(time_per_user, 3600)
#             minutes = remainder / 60

#             # Format the time for each book and add data to the dictionary
#             if hours >= 1:
#                 formatted_time = "{:.0f} h {:.0f} min".format(hours, minutes)
#             else:
#                 formatted_time = "{:.0f} min".format(minutes)

#             formatted_data_per_book[book_id] = {
#                 'formatted_time': formatted_time,
#                 'book_object': book_object,
#             }

#         # Sort the dictionary by formatted time in descending order
#         sorted_data_desc = dict(sorted(formatted_data_per_book.items(), key=lambda item: int(item[1]['formatted_time'].split()[0]), reverse=True))

#         return render(request, 'organizations/average-time-reading-books.html', {'formatted_data_per_book': sorted_data_desc})

class AverageTimeReadingBooks(View):
    def get(self, request):
        # Get user's books and institute name
        user_books = request.user.books.all()
        institute_name = request.user.institute  # Default to 'guest' if not provided

        # Fetch book visits for the given institute
        book_visits = BookVisit.objects.filter(institute_name=institute_name)

        # Use defaultdict to store total time and unique visit count for each book
        total_time_per_book = defaultdict(int)
        unique_visit_count_per_book = defaultdict(set)
        for visit in book_visits:
            if visit.start_time is not None and visit.end_time is not None:
                total_time_per_book[visit.books.id] += (visit.end_time - visit.start_time).seconds
                unique_visit_count_per_book[visit.books.id].add(visit.user.id)

        # Calculate average time per book in hours and minutes
        average_time_per_book = {}
        for book_id, total_time in total_time_per_book.items():
            visit_count = len(unique_visit_count_per_book[book_id])
            if visit_count > 0:
                average_time_seconds = total_time / visit_count
                average_time_hours = int(average_time_seconds // 3600)
                average_time_minutes = int((average_time_seconds % 3600) // 60)
                average_time_per_book[book_id] = {'hours': average_time_hours, 'minutes': average_time_minutes}

        # Render the result
        context = {
            'user_books_and_times': zip(user_books, average_time_per_book.items()),
            'institute_name': institute_name,'active16':'active20'
        }
        return render(request, 'organizations/average-time-reading-books.html', context,)
        

class DetailsCorporate(View):
    def get(self, request):
        data = User.objects.filter(is_organization=True).annotate(book_count=Count('books'),
        student_count=Coalesce(Subquery(User.objects.filter(added_user=OuterRef('id')).values('added_user').annotate(count=Count('id')).values('count')[:1]),Value(0)))
        filter_corporate = User.objects.filter(is_organization=True)
        filter_non_corporate = User.objects.filter(is_organization=False)

        filtering_options = {
            'all': 'All',
            'corporate': 'Corporate',
            'non_corporate': 'Non-Corporate',
        }
        return render(request, 'organizations/details_corporatepage.html', {
            'data': data,
            'active23': 'active12',
            'filter_corporate': filter_corporate,
            'filter_non_corporate': filter_non_corporate,
            'filtering_options': filtering_options,
            
        })

class DeleteCorporate(View):
    def get(self,request,id):
        data = User.objects.get(id=id)
        data.delete()
        return redirect('Detail_organizations')


# class EditCorporate(View):
#     def get(self,request,id):
#         data = User.objects.get(id=id)
#         return render(request,'organizations/edit_corporate.html',{'data':data})
#     def post (self,request,id):
#         data = User.objects.get(id=id)
#         first_name =request.POST.get('first_name')
#         last_name = request.POST.get('last_name')
#         email = request.POST.get('email')
#         phone_number = request.POST.get('phone_number')
#         institute = request.POST.get('institute')
#         User(id=id,first_name=first_name,last_name=last_name,email=email,phone_number=phone_number,institute=institute)
#         data.save()
#         return redirect('Detail_organizations')

class EditCorporate(View):
    def get(self, request, id):
        user = User.objects.get(id=id)
        books = Books.objects.filter(is_active=True)
        return render(request, 'organizations/edit_corporate.html', {'user': user, 'books': books})

    def post(self, request, id):
        user = User.objects.get(id=id)
        institute = request.POST.get('institute')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        phone_number = request.POST.get('phone_number')
        selected_books = request.POST.getlist('books')

        # Update user information
        user.institute = institute
        user.first_name = first_name
        user.last_name = last_name
        user.email = email
        user.phone_number = phone_number
        user.save()

        # Clear existing books for the user
        user.books.clear()

        # Add selected books for the user
        for book_id in selected_books:
            selected_book = Books.objects.get(id=book_id)
            selected_book.is_organization = True
            selected_book.save()
            user.books.add(selected_book)

        return redirect('Detail_organizations')
def organization_login(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        passwords = request.POST.get('password')
        user = authenticate(request=request, email=email, password=passwords)
        if user is not None:
            if User.objects.filter(email=email, is_organization=True):
                dj_login(request, user)
                return redirect('Dashboardorgani')
            else:
                messages.warning(request, 'You Are Not Admin User')
        else:
            messages.warning(request, 'Invalid Email or Password')
    return render(request, 'login.html')

def average_time_spent():
    # Get all user sessions
    all_login_times = ManageLoginTime.objects.filter(
        login_start_time__isnull=False,
        login_end_time__isnull=False,
        institute__isnull = False 
    )
    unique_users = ManageLoginTime.objects.filter(
        login_start_time__isnull=False,
        login_end_time__isnull=False,
        institute__isnull = False 
    ).values('users').distinct().count()
 
    # Calculate the total time spent by adding up the durations of all sessions
    total_duration = timedelta()

    for time in all_login_times:
        # Convert time objects to datetime objects
        start_datetime = datetime.combine(datetime.today(), time.login_start_time)
        end_datetime = datetime.combine(datetime.today(), time.login_end_time)

        session_duration = end_datetime - start_datetime
        total_duration += session_duration

    # Check if there are sessions before calculating averages
    if all_login_times.exists():
        # Calculate the total number of sessions
        # Calculate the overall average time spent per session
        overall_average_time = total_duration / unique_users

        # Convert the overall average time to a human-readable format (optional)
        overall_average_hours, remainder = divmod(overall_average_time.seconds, 3600)
        overall_average_minutes = remainder // 60

        return overall_average_hours, overall_average_minutes
    return 0, 0 

def average_time_spent_admin_dashboard():
    # Get all user sessions
    all_login_times = ManageLoginTime.objects.filter(
        login_start_time__isnull=False,
        login_end_time__isnull=False
    )
    unique_users = ManageLoginTime.objects.filter(
        login_start_time__isnull=False,
        login_end_time__isnull=False
    ).values('users').distinct().count()
 
    # Calculate the total time spent by adding up the durations of all sessions
    total_duration = timedelta()

    for time in all_login_times:
        # Convert time objects to datetime objects
        start_datetime = datetime.combine(datetime.today(), time.login_start_time)
        end_datetime = datetime.combine(datetime.today(), time.login_end_time)

        session_duration = end_datetime - start_datetime
        total_duration += session_duration

    # Check if there are sessions before calculating averages
    if all_login_times.exists():
        # Calculate the total number of sessions
        # Calculate the overall average time spent per session
        overall_average_time = total_duration / unique_users

        # Convert the overall average time to a human-readable format (optional)
        overall_average_hours, remainder = divmod(overall_average_time.seconds, 3600)
        overall_average_minutes = remainder // 60

        return overall_average_hours, overall_average_minutes
    return 0, 0 

class OrganizationsDashboard(View):
    def get(self, request):
        if request.user.is_authenticated:
            if request.user.is_superuser:
                return redirect('dashboard')
        if request.user.is_organization:
            user_id = request.user.id       
            institute = request.user.institute     
            total_student = User.objects.filter(added_user=user_id, institute= institute).count()

            total_unconfirmed = User.objects.filter(added_user=user_id, is_student=False, institute=institute).count()
            total_confirmed = User.objects.filter(added_user=user_id, is_student=True, institute= institute ).count()
            ##### confirmed & unconfirmed graph 
            year = current_year = datetime.now().year
            print(year,'year')
            all_months_years = User.objects.annotate(
            month=ExtractMonth('created_at'),
            year=ExtractYear('created_at')
            ).values('year', 'month').distinct()
            users_month_and_year = all_months_years.annotate(
            unconfirmed=Coalesce(Count('id',filter=Q(added_user=user_id, is_student=False, institute=institute),distinct=True), Value(0)),
            confirmed=Coalesce(Count('id', filter=Q(added_user=user_id, is_student=True, institute= institute,), distinct=True), Value(0))
            )

            all_book = Books.objects.filter(is_active=True).count()
            total_books =  request.user.books.all()
            total_books = len(total_books)

            last_30_days_start = timezone.now() - timedelta(days=30)
            last_3_months_start = timezone.now() - timedelta(days=90)
            total_page_view = BookVisit.objects.filter(institute_name=institute).count()
            total_duration = BookVisit.objects.filter(institute_name=institute).aggregate(
            total_duration=Sum(F('end_time') - F('start_time'))
            )['total_duration']
            if total_duration is not None:
                # Now, total_duration is already a timedelta object
                hours, remainder = divmod(total_duration.seconds, 3600)
                minutes, _ = divmod(remainder, 60)
                total_time = f"Hours: {hours}, Minutes: {minutes}"
            else:
                # If total_duration is None, set default values
                hours, minutes = 0, 0
                total_time = f"Hours: {hours}, Minutes: {minutes}"
                        
            # Get total active users in the last 3 months
            total_active_users_last_3_months = User.objects.filter(
                added_user=request.user.id,
                is_student=True,
                institute = request.user.institute,
                last_login__gte=last_3_months_start
            ).distinct().count()

            # Get total active monthly users in the last 30 days
            total_active_monthly_users = User.objects.filter(
                added_user=request.user.id,
                is_student=True,
                institute = request.user.institute,
                last_login__gte=last_30_days_start
            ).distinct().count()

            # year = current_year = datetime.now().year
            # print(year,'year')

            users_3month_and_30days = User.objects.annotate(
            month=ExtractMonth('created_at'),
            year=ExtractYear('created_at')
            ).values('year', 'month').distinct()
            users_3month_and_30days = all_months_years.annotate(
            total_3month=Coalesce(Count('id',filter=Q(added_user=request.user.id,is_student=True,
            institute = request.user.institute,last_login__gte=last_3_months_start),distinct=True), Value(0)),
            
            total_30days=Coalesce(Count('id', filter=Q(added_user=request.user.id,is_student=True,
            institute = request.user.institute,last_login__gte=last_30_days_start), distinct=True), Value(0))
            )
            # print(users_3month_and_30days,'all_user_30days_3months')



            h , m = average_time_spent()
            student_avg_time = f"Hours: {h}, Minutes: {m}"
            top_students = BookVisit.objects.filter(
            institute_name=request.user.institute,
            user__is_student=True
            ).values('user').annotate(
                total_duration=Sum(F('end_time') - F('start_time'))
            ).order_by('-total_duration')[:10]

            user_objects = []
            for entry in top_students:    
                user_id = entry['user']
                total_duration = entry['total_duration']
                user = User.objects.get(id=user_id)
                if total_duration is not None:
                    hours, remainder = divmod(total_duration.seconds, 3600)
                    minutes, _ = divmod(remainder, 60)
                else:
                    hours, minutes = 0, 0
                user_objects.append({'user': user, 'hours': hours, 'minutes': minutes})
            context = {
                'total_student': total_student,
                'total_Uncondfirmed_Students':total_unconfirmed,
                'total_confirmed_students':total_confirmed,
                'total_books': total_books,
                'all_book':all_book,
                'total_time_reading':total_time,
                'total_page_view':total_page_view,
                'total_active_users_last_3_months':total_active_users_last_3_months,
                'total_active_monthly_users':total_active_monthly_users,
                'student_avg_time':student_avg_time,
                'active11':'active20',
                'top_students':user_objects,
                'year':year,
                'users_month_and_year':users_month_and_year,
                'users_3month_and_30days':users_3month_and_30days
                
            }
            return render(request, 'organizations/organization_dashboard.html',context )



from ebooklib import epub
class BooksStats(View):
    def get(self, request):
        user_book_visits = BookVisit.objects.filter(books__in=request.user.books.all()).values('books').distinct()
        book_list = []
        for book in user_book_visits:
            book_id = book['books']
            book_obj = Books.objects.get(id=book['books'])
            total_page_view = BookVisit.objects.filter(books=book_id).count() or 0
            user_book_visits = BookVisit.objects.filter(books=book_id).values('user').distinct().count() or 0
            d = {
                'book':book_obj,
                'total_page_view':total_page_view,
                'user_book_visits':user_book_visits
            }
            book_list.append(d)
        return render(request, 'organizations/book-status.html', {'book_list':book_list})



class TopStudents(View):
    def get(self, request):
        top_students = BookVisit.objects.filter(
            institute_name=request.user.institute,
            user__is_student=True
        ).values('user').annotate(
            total_duration=Sum(F('end_time') - F('start_time'))
        ).order_by('-total_duration')[:10]

        user_objects = []
        for entry in top_students:    
            user_id = entry['user']
            total_duration = entry['total_duration']

            user = User.objects.get(id=user_id)
            if total_duration is not None:
                hours, remainder = divmod(total_duration.seconds, 3600)
                minutes, _ = divmod(remainder, 60)
            else:
                hours, minutes = 0, 0
            user_objects.append({'user': user, 'hours': hours, 'minutes': minutes})

        return render(request, 'organizations/top-student-lists.html', {'top_students':user_objects} )


# institute_name = getattr(request.user, 'institute', 'guest')
        # book_visits = BookVisit.objects.filter(institute_name=institute_name)

        # total_time_per_book = defaultdict(int)
        # visits_per_book = defaultdict(int)
        # unique_visits_per_book = defaultdict(set)

        # for visit in book_visits:
        #     book = visit.books
        #     visits_per_book[book] += 1
        #     unique_visits_per_book[book].add(visit.user.id)

        # # Process the top N books (you can change this based on your needs)
        # num_books_to_process = 5
        # top_books = sorted(visits_per_book.items(), key=lambda x: x[1], reverse=True)[:num_books_to_process]

        # book_stats = []
        # for book, num_visits in top_books:
        #     total_time = total_time_per_book[book]
        #     unique_visits = len(unique_visits_per_book[book])
        #     avg_time_per_visit = total_time / num_visits if num_visits > 0 else 0

        #     book_stats.append({
        #         'book': book,
        #         'num_visits': num_visits,
        
        #         'unique_visits': unique_visits,
        #         'avg_time_per_visit': avg_time_per_visit,
        #     })

#######  import 
# def create_students_xlsx(request):
#     workbook = Workbook()
#     sheet = workbook.active
#     id = request.user.id
#     added_user_label = f"(User ID: {id})"
#     headers = ["email", "institute","first_name","last_name","phone_number", "is_student",  "added_user", added_user_label]
#     sheet.append(headers)
#     return workbook


def create_students_xlsx():
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "first_name"
    sheet["B1"] = "last_name"
    sheet["C1"] = "email"
    sheet["D1"] = "phone_number"
    return workbook



class StudentsCreateXLSX(View):
    def get(self, request):
        if request.user.is_organization:
            workbook = create_students_xlsx()
            if workbook is not None:  # Check if the workbook is not None
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=students.xlsx'
                virtual_workbook = BytesIO()
                workbook.save(virtual_workbook)
                virtual_workbook.seek(0)
                response.write(virtual_workbook.getvalue())
                return response
        return redirect('students_details')




class StudentsUpload(View):
    def get(self, request):
        user_id = request.user.id
        return render(request, 'organizations/student_import.html')
    def post(self, request):
        user_id = request.user.id
        if not request.user.is_organization:
            messages.error(request, 'Unauthorized access')
            return redirect('students_details')
        emp_resource = UserResource()
        new_student = request.FILES.get('myfile')
        dataset = tablib.Dataset()
        dataset.load(new_student.read(), format='xlsx')
        # Check if email already exists before importing
        existing_emails = set(User.objects.values_list('email', flat=True))
        new_emails = set(row.get('email') for row in dataset.dict)
        duplicate_emails = existing_emails.intersection(new_emails)
        if duplicate_emails:
            messages.error(request,f'Email id  {duplicate_emails} already exists. Please use a different email to create an account.')
            return redirect('students_details')
        result = emp_resource.import_data(dataset, dry_run=True)
        if not result.has_errors():
            emp_resource.import_data(dataset, dry_run=False)
            # Send email to each student
            institute = request.user.institute
            for row in dataset.dict:
                student_email = row.get('email')
                try:
                    stu = User.objects.get(email=student_email)
                except ObjectDoesNotExist:
                    messages.warning(request, f"User with email {student_email} not found.")
                    continue
                stu.institute = institute
                stu.added_user = request.user.id
                stu.is_active = False
                stu.save()
                student_id = stu.id
                print("student_email",student_email)
                if student_email:
                    subject = "Congratulations!"
                    link = f"http://64.227.157.173:8002/Admin/students_registration/{student_id}"
                    message = f"You have been invited by {institute} to join LawAfrica E-Book Platform. {link}"
                    from_email = "indy@optimuse-solutions.com"
                    recipient_list = [student_email]
                    email = EmailMessage(subject, message, from_email, recipient_list)
                    email.send()
                    messages.success(request,'Bulk Data Uploaded Successfully')
                    return redirect('students_details')
            messages.error(request, 'Blank file data upload unsuccessful ')
            return redirect('students_details')




class StudentCreate(View):
    def get(self, request):
        if request.user.is_organization:
            user = User.objects.filter(added_user=request.user.id, is_student=True)
            return render(request, 'organizations/create_student.html', {'user': user})
    def post(self, request):
        if request.user.is_organization:
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            email = request.POST.get('email')
            phone_number = request.POST.get('phone_number')
            password = request.POST.get('password')
            institute = request.POST.get('institute')
            if User.objects.filter(email=email, added_user=request.user.id).exists():
                messages.error(request,'Email id already exists. Please use a different email to create an account.')
            else:
                user = User.objects.create_user(
                    password=password,
                    institute=institute,
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                    phone_number=phone_number,
                )
                
                user.added_user = request.user.id 
                user.is_organization= False 
                user.save()
                email = user.email
                subject = "Congratulations!"
                link =" http://64.227.157.173:8002/Admin/students_registration/"+str(user.id)
                message = f"You have been invited by {institute} to join LawAfrica E-Book Platform. " + link

                # Render the email template with message details
                email_body = render_to_string('organizations/send_email.html', {
                    'subject': subject,
                    'message': message,
                    'id': user.id,
                    'link':link,
                    'email':email
                    
                })
                from_email = "indy@optimuse-solutions.com"  # Replace with your email
                recipient_list = [email]
                email = EmailMessage(subject, email_body, from_email, recipient_list)
                email.content_subtype = "html"  # Set the content type to HTML
                email.send()
                messages.success(request, 'Student created successfully.')
            return redirect('students_details')



class StudentsRegister(View):
    def get(self, request, id):
        user = User.objects.get(id=id)
        user.save()
        added_user = user.added_user if user.added_user else None
        print(added_user, 'added_user_id')
        email = user.email
        first_name = user.first_name
        last_name = user.last_name
        phone_number = user.phone_number
        institute = user.institute
        return render(request, 'organizations/student_registration.html', {
            'email': email,
            'first_name': first_name,
            'last_name': last_name,
            'phone_number': phone_number,
            'added_user_id': added_user,
        })

    def post(self, request, id):
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        phone_number = request.POST.get('phone_number')
        password = request.POST.get('password')
        # Get the user who added the current student
        user = User.objects.get(id=id)
        institute = user.institute
        added_user = user.added_user if user.added_user else None
        # Create a new student instance with the desired attributes
        new_student = User(
            id=id,
            first_name=first_name,
            last_name=last_name,
            email=email,
            phone_number=phone_number,
            institute=institute,
            is_student=True,
            added_user=added_user,
        )
        # Set the password using set_password method to properly hash it
        new_student.set_password(password)
        new_student.save()
        activation_link = request.build_absolute_uri(reverse('activate_account', args=[new_student.id]))
        subject = 'Activate Your Account'
        message = f'Hi {new_student.first_name},\n\nPlease click on the following link to activate your account:\n\n{activation_link}'
        from_email = 'indy@optimuse-solutions.com'  
        send_mail(subject, message, from_email, [new_student.email])
        messages.success(request, 'Student registration successfully. Please check your email to activate your account.')
        return redirect('enduser-login')



####### students login 
# def students_login(request):
#     if request.method == 'POST':
#         email = request.POST.get('email')
#         password = request.POST.get('password')
#         remember_me = request.POST.get('remember_me')
#         user = authenticate(request=request, email=email, password=password)
#         if user is not None:
#             if User.objects.filter(email=email, is_student=True):
#                 dj_login(request, user)
#                 if remember_me: 
#                     request.session.set_expiry(0)  
#                 else:
#                     request.session.set_expiry(None)
#                 return redirect('Customer_dashboard')
#             elif User.objects.filter(email=email, is_student=True):
#                 dj_login(request, user)
#                 if remember_me:
#                     request.session.set_expiry(0)
#                 else:
#                     request.session.set_expiry(None)
#                 return redirect('Customer_dashboard')
#             else:
#                 messages.warning(request, 'You Are Not Students ')
#         else:
#             messages.warning(request, 'Invalid Email or Password')
#     return render(request, 'organizations/student_login.html')



def students_login(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        # Authenticate the user
        user = authenticate(request=request, email=email, password=password)

        if user is not None and user.is_authenticated:
            # Check if the user is a valid student and belongs to the organization
            if user.is_student and user.added_user == request.user:
                # Use Django's login function to persist the user in the session
                login(request, user)
                login_time = datetime.now().time()
                messages.success(request, 'Login successfully.')
                return redirect('Customer_dashboard')
            else:
                messages.warning(request, 'Invalid Email or Password or You Are Not a Valid Student for this Organization.')
        else:
            messages.warning(request, 'Invalid Email or Password.')

    return render(request, 'organizations/student_login.html')



# email_to_delete = "mobappssolutions149@gmail.com"  
# User.objects.filter(email=email_to_delete).delete()



class StudentsDetails(View):
    def get(self, request):
        if request.user.is_authenticated and request.user.is_organization:
            status_filter = request.GET.get('status_filter', 'all')
            students = User.objects.filter(added_user=request.user.id, institute=request.user.institute)

            if status_filter == 'confirmed':
                students = students.filter(added_user=request.user.id,is_student=True)

            elif status_filter == 'unconfirmed':
                students = students.filter(added_user=request.user.id,is_student=False, is_organization=False)

            return render(request, 'organizations/students_details.html', {'students': students, 'status_filter': status_filter,'active17':'active20'})
        
        return redirect('students_create')



class EditStudents(View):
    def get(self, request, id):
        data = User.objects.get(id=id)
        return render(request, 'organizations/edit_students.html', {'data': data})
    def post(self, request, id):
        data = User.objects.get(id=id)
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        phone_number = request.POST.get('phone_number')
        data.first_name = first_name
        data.last_name = last_name
        data.email = email
        data.phone_number = phone_number
        data.save()
        return redirect('students_details')


class DeleteStudent(View):
    def get(self,request,id):
        data = User.objects.get(id=id)
        data.delete()
        return redirect('students_details')
    


class CorporateUpdateprofile(View):
    def get(self, request):
        if request.user.is_authenticated:
            data=User.objects.filter(id=request.user.id,is_organization=True)
            return render(request, 'organizations/profile_organization.html',{'data':data,'active18':'active20'})
        return redirect('login')
    def post(self, request):
        id = request.user.id
        usern = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        phone_number = request.POST.get('phone_number')
        profile_pic = request.FILES.get('profile_pic')
        if profile_pic is not None:
            id = request.user.id
            user = User.objects.get(id=id)
            user.first_name = usern
            user.email = email
            user.last_name = last_name
            user.phone_number = phone_number
            user.profile_pic = profile_pic
            user.save()
            messages.success(request,'User Profile Updated Successfully!')
            return redirect('corporate_profile')
        else:
            user = User.objects.get(id=id)
            user.first_name = usern
            user.email = email
            user.last_name = last_name
            user.phone_number = phone_number
            user.profile_pic = user.profile_pic
            user.save()
            messages.success(request,'User Profile Updated Successfully!!')
            return redirect('corporate_profile')
        

class OrganizationChangePassword(View):
    def post(self, request):
        old_password = request.POST.get('old_password')
        password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        if old_password and password and confirm_password:
            if request.user.check_password(old_password):
                if password == confirm_password:
                    user = User.objects.get(id=request.user.id)
                    user.set_password(password)
                    user.save()
                    update_session_auth_hash(request, user)
                    messages.success(request, 'Password changed successfully.')
                    return redirect('Dashboardorgani')  
                else:
                    messages.error(request, 'New password and confirm password do not match.')
            else:
                messages.error(request, 'Old password is incorrect.')
        else:
            messages.error(request, 'All fields are required.')

        return redirect('corporate_profile')
    




class CorporateAddCollection(View):
    def get(self, request):
        data = Books.objects.all()
        return render(request, 'organizations/Addcollection_corporate.html', {'data': data})
    def post(self, request):
        name = request.POST.get('name')
        image = request.FILES.get('image')
        books = request.POST.getlist('books')
        data = Collections.objects.create(user=request.user,name=name,image=image)
        for book_id in books:
            book = Books.objects.get(id=book_id)
            CollectionsList.objects.create(user=request.user,Collections_name=data,books_id=book)
        return redirect('listcollectioncorp')
    

from itertools import groupby

class CorporateLibrary(View):
    def get(self, request):
        user_books = request.user.books.all().order_by('-id')
        st = request.GET.get('searchname')

        # Apply search filter if 'st' is not None
        if st:
            user_books = user_books.filter(
                Q(title__icontains=st) |
                Q(author__icontains=st) |
                Q(publication_date__icontains=st) |
                Q(price__icontains=st) |
                Q(category__book_category__icontains=st)
            )

        # Group books by category
        grouped_books = {key: list(group) for key, group in groupby(user_books, key=lambda x: x.category.book_category)}
        return render(request, 'organizations/corporate_booklist.html', {'grouped_books': grouped_books,'active13':'active20'})


class CorporateCollectionList(View):
    def get(self, request):
        data = Collections.objects.filter(user=request.user)
        return render(request,'organizations/detailscollection_corporate.html', {'data': data})


class CorporateCollectionView(View):
    def get(self,request,id):
        data = CollectionsList.objects.filter(user=request.user,Collections_name=Collections.objects.get(id=id))
        return render(request,'organizations/collection_view.html',{'data':data})
    



class CorporateCollectionViewDelete(View):
    def get(self,request,id):
        data = Collections.objects.get(user=request.user,id=id)
        data.delete()
        return redirect('listcollectioncorp')
    
class CreateSurvey(View):
    def get(self,request):
        data = SurveyResponse.objects.all()
        return render(request,'create_survey.html',{'data':data, 'active29':'active12'})
    def post(self, request):
        user = request.user

        # Get all the submitted data from the form
        questions = request.POST.getlist('question')
        answers1 = request.POST.getlist('answer1')
        answers2 = request.POST.getlist('answer2')
        answers3 = request.POST.getlist('answer3')

        # Zip the lists to iterate over them together
        survey_data = zip(questions, answers1, answers2, answers3)

        # Create SurveyResponse objects for each set of data
        for question, answer1, answer2, answer3 in survey_data:
            SurveyResponse.objects.create(
                user=user,
                question=question,
                answer1=answer1,
                answer2=answer2,
                answer3=answer3
            )
        return redirect('UserservayDetails')



    

from django.db.models import Count
from django.db.models import Max


class CustomerServeyDetail(View):
    def get(self, request):
        latest_responses = SurveyResponse.objects.filter(
            user__is_superuser=False
        ).values('user__email').annotate(max_id=Max('id'))

        data = SurveyResponse.objects.filter(
            user__is_superuser=False, 
            id__in=[entry['max_id'] for entry in latest_responses]
        ).values('user__email', 'id','created_at').distinct()

        return render(request, 'survery_deatils.html', {'data': data})


def survey(request):
    return render(request,'backend_cmspage/survey.html')

class ViewStudentsServey(View):
    def get(self,request,id):
        data = SurveyResponse.objects.filter(user__is_superuser=False)
        return render(request,'view_servey_students.html',{'data':data})


class CustomerservayDelete(View):
    def get(self,request,id):
        data = SurveyResponse.objects.get(id=id)
        data.delete()
        return redirect('CustomerServeyDetail')





class UserservayDelete(View):
    def get(self,request,id):
        data = SurveyResponse.objects.get(id=id)
        data.delete()
        return redirect('UserservayDetails')



# class StudentsUpload(View):
#     def get(self, request):
#         return render(request, 'organizations/student_import.html')
#     def post(self, request):
#         if request.user.is_organization:
#             emp_resource = UserResource()
#             new_employee = request.FILES.get('myfile')
#             if new_employee:
#                 dataset = tablib.Dataset()
#                 dataset.load(new_employee.read(), format='xlsx')
#                 result = emp_resource.import_data(dataset, dry_run=True)
#                 fields = dataset.headers
#                 if not result.has_errors():
#                     emp_resource.import_data(dataset, dry_run=False)
#                     # Send email to each student
#                     for row in dataset.dict:
#                         student_email = row.get('email')
#                         student_id = User.objects.filter(email=student_email)
#                         institute = row.get('institute')
#                         if student_email:
#                             subject = "Congratulations!"
#                             message = f"You have been invited by {institute} to join LawAfrica E-Book Platform."

#                             # Render the email template with message details
#                             email_body = render_to_string('organizations/send_email.html', {
#                                 'subject': subject,
#                                 'message': message,
#                                 'student_id':student_id,

                                
#                             })
#                             from_email = "your_email@gmail.com"  # Replace with your email
#                             recipient_list = [student_email]
#                             email = EmailMessage(subject, email_body, from_email, recipient_list)
#                             email.content_subtype = "html"  # Set the content type to HTML
#                             email.send()
#                     messages.success(request, 'Bulk Data Updated Successfully')
#                     return redirect('students_details')
#                 else:
#                     error_messages = [e for e in result.row_errors()]
#                     messages.error(request, f'Data Upload Unsuccessful: {error_messages}')
#             else:
#                 messages.error(request, 'No file uploaded.')
#         else:
#             messages.error(request, 'Unauthorized access')
#         return redirect('students_create')










#####
def home(request):
    return render(request,'backend_cmspage/home.html')


def blogs(request):
    return render(request,'backend_cmspage/blogs.html')


def digital_publishing(request):
    return render(request,'backend_cmspage/digital_publishing.html')

def ebook_store(request):
    return render(request,'backend_cmspage/e-book_store.html')



def indybot(request):
    return render(request,'backend_cmspage/indybot.html')


def contactus(request):
    return render(request,'backend_cmspage/contact_us.html')



def add_subscription(request):
    if request.method == 'POST':
        user = request.user 
        start_date = request.POST.get('start_date')
        duration_months = int(request.POST.get('duration_months'))
        expiration_reminder = request.POST.get('expiration_reminder')
        if duration_months not in [6, 12]:
            pass
        else:
            subscription = Subscription.objects.create(
                user=user,
                start_date=start_date,
                duration_months=duration_months,
                expiration_reminder=expiration_reminder
            )
            if expiration_reminder:
                # Calculate the reminder date (1 month before end_date)
                reminder_date = subscription.end_date - relativedelta(months=1)
                if datetime.now().date() >= reminder_date:
                    send_mail(
                        'Subscription Expiration Reminder',
                        'Your subscription will expire in one month.',
                        'sender@example.com',
                        [user.email],
                        fail_silently=False,
                    )
            return redirect('subscription_success')  
    return render(request, 'add_subscription.html')



class UserservayDetails(View):
    def get(self,request):
        data = SurveyResponse.objects.filter(user__is_superuser=True)
        return render(request,'user_servaydetails.html',{'data':data})
    


class UserservayDelete(View):
    def get(self,request,id):
        data = SurveyResponse.objects.get(id=id)
        data.delete()
        return redirect('UserservayDetails')
    



# class CorporateLibrary(View):
#     def get(self, request):
#         user = request.user
#         user_books = request.user.books.all()
#         st = request.GET.get('searchname')

#         # Apply search filter if 'st' is not None
#         if st:
#             user_books = user_books.filter(
#                 Q(title__icontains=st) |
#                 Q(author__icontains=st) |
#                 Q(publication_date__icontains=st) |
#                 Q(price__icontains=st) |
#                 Q(category__book_category__icontains=st)
#             )
#         grouped_books = {key: list(group) for key, group in groupby(user_books, key=lambda x: x.category.book_category)}
#         return render(request, 'organizations/corporate_booklist.html', {'grouped_books': grouped_books})


class CorporateAllBooks(View):
    def get(self,request):
        user = request.user
        all_books = request.user.books.all()
        st = request.GET.get('searchname')
    
        # Apply search filter if 'st' is not None
        if st:
            all_books = all_books.filter(
                Q(title__icontains=st) |
                Q(author__icontains=st) |
                Q(publication_date__icontains=st) |
                Q(price__icontains=st) |
                Q(category__book_category__icontains=st)
            )
        return render(request,'organizations/corporate_all_book.html',{'all_books':all_books,'active12':'active20'})
    

class DeleteMultipleUser(View):
    def get(self, request, id):
        data = User.objects.get(id=id)
        data.delete()
        return redirect('user_details')
    def post(self, request):
        # Handle POST request for deleting multiple users
        user_ids = request.POST.getlist('userIDs[]')

        # Perform deletion of users with IDs in user_ids
        User.objects.filter(id__in=user_ids).delete()
        messages.success(request,'Data deleted successfully')
        return redirect('user_details')


class DeleteMultipleBooks(View):
    def get(self, request, id):
        data = Books.objects.get(id=id)
        data.delete()
        return redirect('details_book')
    def post(self, request):
        # Handle POST request for deleting multiple users
        book_ids = request.POST.getlist('bookIDs[]')

        # Perform deletion of users with IDs in user_ids
        Books.objects.filter(id__in=book_ids).delete()
        messages.success(request,'Data deleted successfully')
        return redirect('details_book')
    

class StudentsListConve(View):
    def get(self,request):
        data = User.objects.filter(is_student=True)
        return render(request,'student_conversion_list.html',{'data':data})
    
class DeleteStudentListConve(View):
    def get(self,request,id):
        data = User.objects.get(id=id)
        data.delete()
        return redirect('StudentsListConve')

class ActiveUserList(View):
    def get(self, request):
        time_range = request.GET.get('time_range', 'daily')
        today = timezone.now()

        if time_range == 'daily':
            start_date = today - timedelta(days=1)
        elif time_range == 'weekly':
            start_date = today - timedelta(weeks=1)
        elif time_range == 'monthly':
            start_date = today - timedelta(weeks=4)
        elif time_range == 'last3months':
            start_date = today - timedelta(weeks=12)
        else:
            start_date = today - timedelta(days=1)

        data = User.objects.filter(is_active=True).exclude(is_superuser=True)
        return render(request, 'active_user.html', {'data': data, 'time_range': time_range})



class ActiveUserDelete(View):
    def get(self,request,id):
        data = User.objects.get(id=id)
        data.delete()
        return redirect('ActiveUserList')


class UserChangePassword(View):
    def get(self, request,id):
        data = User.objects.get(id=id)
        return render(request,'change_password.html',{'data':data})
    def post(self, request,id):
        email = request.POST.get('email')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        if email and new_password and confirm_password:
            if new_password == confirm_password:
                if User.objects.filter(email=email).exists():
                    new_password = confirm_password
                    user = User.objects.get(email=email)
                    user.set_password(new_password)
                    user.save()
                    messages.success(request,'Password Changed Successfully.!!!')
                    return redirect('user_details')
                else:
                    messages.error(request, 'Email id is not exist')
                return redirect('ChangePassword',id=id)
            else:
                messages.error(request,'Please check your new password and confirm password..!!')
        else:
            messages.error(request,'Email and new_password and confirm_password is required')
        return redirect('changepassword',id=id)


class BookAssign(View):
    def get(self, request, id):
        user = User.objects.get(id=id)
        saved_books = user.books.values_list('id', flat=True)  # Get IDs of saved books
        data = Books.objects.filter(is_active=True)
        select_categories = BookCategory.objects.all()
        return render(request, 'book_assign.html', {'data': data, 'select_categories': select_categories, 'saved_books': saved_books})
    def post(self, request, id):
        user = User.objects.get(id=id)
        selected_books = request.POST.getlist('books')
        for book_id in selected_books:
            selected_book = Books.objects.get(id=book_id)
            selected_book.is_organization = True
            selected_book.save()
            user.books.add(selected_book)
        return redirect('Detail_organizations')